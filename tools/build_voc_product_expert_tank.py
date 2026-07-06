from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import re
import shutil

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


GROUPS = {
    "25-34岁": ["B0F6YG561Y", "B0G2LT7279", "B0GV41HN7Q", "B0GH8D9T3P", "B0G4W38Q54"],
    "35-44+": ["B0FC2LHG2M", "B0FDF3DKFZ", "B0G39N131T", "B0GS976T5W"],
}
SOURCE_DIRS = [
    Path("G:/伊羽霓/AI-开发/20260624背心/日常背心"),
    Path("日常背心"),
    Path("outputs/web_report_jobs/20260703_092715/uploads"),
]
OUT_DIR = Path("outputs/20260703_voc_product_expert")
SOURCE_COPY = OUT_DIR / "source_inputs"
OUT_FILE = OUT_DIR / f"VOC_PRODUCT_背心产品专家分析_{datetime.now().strftime('%H%M%S')}.xlsx"


TAXONOMY = [
    {
        "label": "贴肤舒适/无勒痕无刺激",
        "pos": ["comfortable", "comfy", "soft", "buttery", "smooth", "second skin", "no digging", "no marks", "stretchy"],
        "neg": ["uncomfortable", "itchy", "scratchy", "irritation", "rash", "red marks", "dig", "digs", "chafing", "tight", "pain"],
        "mechanism": "面料触感、边口压力、内置杯/下围接触面共同决定贴肤体验。",
        "kano": "基本型",
    },
    {
        "label": "内置罩杯/免穿内衣",
        "pos": ["built in bra", "built-in bra", "shelf bra", "no bra", "bra-free", "padded bra", "built in cups"],
        "neg": ["bra design", "cup", "cups", "padding", "pad", "pads", "pad moves", "pads move", "cups move"],
        "mechanism": "杯垫固定方式、杯型厚薄、下围弹力和外层面料决定免内衣体验。",
        "kano": "期望型",
    },
    {
        "label": "支撑承托/DD+适配",
        "pos": ["supportive", "support", "holds", "held", "lift", "large chest", "big chest", "ddd", "dd"],
        "neg": ["no support", "not supportive", "lacks support", "not enough support", "not for large", "falls", "sag"],
        "mechanism": "支撑来自下围、杯宽、侧翼高度、肩带位置和面料模量。",
        "kano": "期望型",
    },
    {
        "label": "遮点/浅色不透安全",
        "pos": ["not see through", "opaque", "not sheer", "double lined", "coverage", "thick enough"],
        "neg": ["see through", "sheer", "thin", "pad through", "pads visible", "nipple", "visible", "white tank"],
        "mechanism": "面料克重、双层结构、杯色匹配和杯边过渡会影响浅色安全感。",
        "kano": "基本型",
    },
    {
        "label": "版型显瘦/胸型自然",
        "pos": ["flattering", "fits well", "nice shape", "cute", "sexy", "snatched", "slimming", "curves", "shape"],
        "neg": ["weird fit", "awkward", "unflattering", "flatten", "uniboob", "spillage", "side boob", "armpit", "boobs"],
        "mechanism": "领型、侧翼、杯宽、腰腹贴合和肩带点位共同决定视觉比例。",
        "kano": "期望型",
    },
    {
        "label": "尺码合身/尺码表准确",
        "pos": ["true to size", "perfect fit", "fits perfectly", "size was perfect", "tts"],
        "neg": ["too small", "too big", "runs small", "runs big", "size up", "not true to size", "size chart"],
        "mechanism": "S/M/L 到胸围、下胸围、杯围和衣长的映射决定试错率。",
        "kano": "基本型",
    },
    {
        "label": "衣长比例/crop长短适中",
        "pos": ["perfect length", "good length", "right length", "cropped", "crop top", "not too short", "not too long"],
        "neg": ["too long", "too short", "not cropped", "very cropped", "super cropped", "crop is too", "length", "longer than", "shorter than"],
        "mechanism": "衣长、下摆位置和胸腰比例决定外穿时显高显瘦，过长会堆积，过短会露肤和上移。",
        "kano": "期望型",
    },
    {
        "label": "面料质量/做工耐用",
        "pos": ["quality", "well made", "nice material", "washes well", "durable", "thick fabric"],
        "neg": ["poor quality", "cheap", "seams", "stitch", "broke", "wash", "fraying", "threads"],
        "mechanism": "缝线、杯垫固定、面料回弹和洗后稳定性决定复购信任。",
        "kano": "基本型",
    },
    {
        "label": "无痕隐形/衣服下平整",
        "pos": ["seamless", "smooth under", "under clothes", "under shirt", "under blazer", "no lines"],
        "neg": ["lines", "indentation", "bulge", "shows under", "not smooth", "visible lines"],
        "mechanism": "边口厚度、杯边过渡、侧缝和面料弹性决定贴身外衣下的平整度。",
        "kano": "期望型",
    },
    {
        "label": "底摆稳定/不卷边不上移",
        "pos": ["stays down", "doesn't ride up", "does not roll", "lays flat"],
        "neg": ["rolls", "roll up", "rolls up", "curls", "bunches", "rides up", "band rolls"],
        "mechanism": "衣长、底摆弹力、腰腹版型、克重和身体曲线适配共同影响卷边。",
        "kano": "基本型",
    },
    {
        "label": "透气/夏季轻薄",
        "pos": ["breathable", "lightweight", "cool", "summer", "not hot"],
        "neg": ["hot", "sweaty", "heavy", "too thick", "not breathable"],
        "mechanism": "杯垫厚度、面料层数、纤维结构和贴身面积影响闷热感。",
        "kano": "期望型",
    },
    {
        "label": "肩带/可调节穿法",
        "pos": ["adjustable straps", "adjustable", "removable straps", "extra straps", "strapless", "convertible"],
        "neg": ["not adjustable", "straps not adjustable", "can't adjust", "cannot adjust", "straps too long", "straps too short", "straps broke", "straps dig", "strap attachment", "strap slips", "straps fall"],
        "mechanism": "肩带长度、弹力、连接点加固和可转换结构决定场景扩展。",
        "kano": "魅力型",
    },
    {
        "label": "外观颜色/时尚性感",
        "pos": ["color", "colors", "pretty", "beautiful", "cute", "stylish", "elegant", "sexy"],
        "neg": ["wrong color", "color off", "not as pictured", "different color", "ugly"],
        "mechanism": "颜色还原、页面图片一致性和领型风格影响点击与外穿信心。",
        "kano": "魅力型",
    },
    {
        "label": "价格/性价比",
        "pos": ["worth it", "great price", "affordable", "value", "buy again"],
        "neg": ["not worth", "overpriced", "waste of money"],
        "mechanism": "价格感知会放大或缓冲质量、舒适和罩杯结构评价。",
        "kano": "魅力型",
    },
]

NEGATIVE_LABELS = {
    "贴肤舒适/无勒痕无刺激": "勒痕/刺痒/不舒适",
    "内置罩杯/免穿内衣": "罩杯结构不稳/杯垫问题",
    "支撑承托/DD+适配": "支撑不足/DD+不适配",
    "遮点/浅色不透安全": "浅色透/露点/杯垫显形",
    "版型显瘦/胸型自然": "胸型怪/压胸/副乳外溢",
    "尺码合身/尺码表准确": "尺码偏差/尺码表不准",
    "衣长比例/crop长短适中": "crop衣长太长/太短",
    "面料质量/做工耐用": "质量差/做工洗后问题",
    "无痕隐形/衣服下平整": "衣服下显线/不平整",
    "底摆稳定/不卷边不上移": "底摆卷边/上移/堆积",
    "透气/夏季轻薄": "闷热/厚重/不透气",
    "肩带/可调节穿法": "肩带勒/断裂/不稳定",
    "外观颜色/时尚性感": "颜色不符/外观不如图",
    "价格/性价比": "不值/价格偏高",
}

POSITIVE_KEYWORD_LABELS = [
    (["true to size", "perfect fit", "fits perfectly", "size was perfect", "tts"], "尺码准确/合身"),
    (["comfortable", "comfy"], "舒适好穿"),
    (["soft", "buttery", "smooth", "second skin"], "面料柔软亲肤"),
    (["supportive", "support", "holds", "held", "lift"], "支撑承托好"),
    (["large chest", "big chest", "ddd", "dd"], "大胸可穿/承托够"),
    (["perfect length", "good length", "right length", "cropped", "crop top", "not too short", "not too long"], "crop衣长合适"),
    (["built in bra", "built-in bra", "shelf bra", "no bra", "bra-free", "padded bra", "built in cups"], "内置罩杯/免穿内衣"),
    (["not see through", "opaque", "not sheer", "double lined", "coverage", "thick enough"], "遮点不透/安全感强"),
    (["flattering", "fits well", "nice shape", "snatched", "slimming", "curves", "shape"], "显瘦显身材"),
    (["cute", "sexy"], "外观好看/可外穿"),
    (["quality", "well made", "nice material", "washes well", "durable", "thick fabric"], "质量做工好"),
    (["seamless", "smooth under", "no lines", "invisible"], "无痕平整"),
    (["doesn't roll", "does not roll", "stays put", "no ride up"], "底摆稳定不卷边"),
    (["breathable", "lightweight", "cool"], "轻薄透气"),
    (["adjustable straps", "strapless", "straps"], "肩带/穿法方便"),
    (["color", "colors", "pretty", "looks like picture"], "颜色外观满意"),
    (["worth it", "great price", "affordable", "value", "buy again"], "性价比高/愿意复购"),
]

POSITIVE_TOP_WORD_LABELS = {
    "comfortable": "舒适好穿",
    "comfy": "舒适好穿",
    "soft": "面料柔软",
    "buttery": "黄油感亲肤面料",
    "smooth": "顺滑亲肤",
    "second skin": "第二层皮肤感",
    "true to size": "尺码准确",
    "perfect fit": "合身度好",
    "fits perfectly": "合身度好",
    "supportive": "支撑好",
    "support": "支撑好",
    "lift": "有提托感",
    "holds": "承托稳定",
    "held": "承托稳定",
    "large chest": "大胸可穿",
    "big chest": "大胸可穿",
    "perfect length": "衣长刚好",
    "good length": "衣长合适",
    "right length": "衣长合适",
    "cropped": "短款比例好",
    "crop top": "短款外穿感好",
    "built in bra": "内置文胸方便",
    "built-in bra": "内置文胸方便",
    "shelf bra": "内置托胸结构方便",
    "no bra": "可免穿内衣",
    "not see through": "不透安全",
    "opaque": "不透安全",
    "double lined": "双层遮点",
    "flattering": "显身材",
    "snatched": "收腰显瘦",
    "quality": "质量好",
    "well made": "做工好",
    "seamless": "无痕",
    "no lines": "衣服下无线条",
    "stays put": "不跑位",
    "breathable": "透气",
    "lightweight": "轻薄",
    "adjustable straps": "肩带可调",
    "pretty": "外观好看",
    "worth it": "值得买",
    "buy again": "愿意复购",
}

NEGATIVE_KEYWORD_LABELS = [
    (["no support", "not supportive", "lacks support", "not enough support"], "无支撑/支撑不足"),
    (["sag", "falls"], "支撑不足/下垂滑落"),
    (["not for large", "large chest", "big chest", "ddd", "dd"], "大胸不适配"),
    (["too small", "runs small", "size up"], "尺码偏小/需拍大"),
    (["too big", "runs big"], "尺码偏大"),
    (["not true to size", "size chart"], "尺码表不准"),
    (["too long", "longer than"], "crop太长"),
    (["too short", "very cropped", "super cropped", "shorter than"], "crop太短"),
    (["not cropped", "crop is too", "length"], "crop衣长不合适"),
    (["see through", "sheer", "thin"], "面料薄/透"),
    (["nipple", "visible", "pads visible", "pad through"], "露点/杯垫显形"),
    (["pad moves", "pads move", "cups move"], "杯垫移位"),
    (["cup", "cups", "padding", "pad", "pads"], "罩杯结构问题"),
    (["uniboob", "flatten"], "压胸/胸型不自然"),
    (["spillage", "side boob", "armpit"], "副乳/腋下外溢"),
    (["weird fit", "awkward", "unflattering"], "版型不合身"),
    (["uncomfortable", "itchy", "scratchy", "irritation", "rash", "red marks", "dig", "digs", "chafing", "tight", "pain"], "勒痕/刺痒/不舒适"),
    (["poor quality", "cheap", "seams", "stitch", "broke", "wash", "fraying", "threads"], "质量差/做工洗后问题"),
    (["rolls", "roll up", "rolls up", "ride up"], "底摆卷边/上移"),
    (["hot", "sweaty", "thick", "not breathable"], "闷热/厚重"),
    (["not adjustable", "straps not adjustable", "can't adjust", "cannot adjust"], "肩带不可调节"),
    (["straps too long", "straps too short"], "肩带长短不合适"),
    (["strap digs", "straps dig", "strap broke", "straps broke", "straps fall", "strap slips"], "肩带勒/断裂/滑落"),
    (["color mismatch", "not as pictured", "fades"], "颜色不符/不如图"),
    (["not worth", "overpriced", "waste of money"], "不值/价格偏高"),
]

NEGATIVE_TOP_WORD_LABELS = {
    "no support": "无支撑",
    "not supportive": "支撑不足",
    "lacks support": "支撑不足",
    "not enough support": "支撑不够",
    "sag": "支撑不足/下垂",
    "falls": "支撑滑落",
    "not for large": "大胸不适配",
    "too small": "尺码偏小",
    "runs small": "版型偏小",
    "size up": "需要拍大一码",
    "too big": "尺码偏大",
    "runs big": "版型偏大",
    "not true to size": "尺码不准",
    "size chart": "尺码表不准",
    "too long": "crop太长",
    "longer than": "crop太长",
    "too short": "crop太短",
    "very cropped": "crop太短",
    "super cropped": "crop太短",
    "shorter than": "crop太短",
    "not cropped": "不像短款/衣长偏长",
    "crop is too": "crop衣长不合适",
    "length": "衣长不合适",
    "see through": "透/不安全",
    "sheer": "透/不安全",
    "thin": "面料薄",
    "nipple": "露点风险",
    "visible": "杯垫/胸点显形",
    "pads visible": "杯垫显形",
    "pad through": "杯垫透出",
    "pad moves": "杯垫移位",
    "pads move": "杯垫移位",
    "cups move": "罩杯移位",
    "padding": "杯垫问题",
    "pad": "杯垫问题",
    "pads": "杯垫问题",
    "cup": "罩杯不合适",
    "cups": "罩杯不合适",
    "uniboob": "压成一片胸",
    "flatten": "压胸",
    "spillage": "胸部外溢",
    "side boob": "侧胸外露",
    "armpit": "腋下副乳明显",
    "weird fit": "版型怪",
    "awkward": "穿着尴尬",
    "unflattering": "不显身材",
    "uncomfortable": "不舒适",
    "itchy": "刺痒",
    "scratchy": "刮皮肤",
    "irritation": "皮肤刺激",
    "rash": "皮肤过敏/红疹",
    "red marks": "勒出红印",
    "dig": "勒肉",
    "digs": "勒肉",
    "chafing": "摩擦不适",
    "tight": "偏紧勒身",
    "pain": "穿着疼痛",
    "poor quality": "质量差",
    "cheap": "廉价感",
    "seams": "缝线问题",
    "stitch": "车线问题",
    "broke": "损坏",
    "wash": "洗后问题",
    "fraying": "脱线毛边",
    "threads": "线头问题",
    "rolls": "底摆卷边",
    "roll up": "底摆卷起",
    "rolls up": "底摆卷起",
    "ride up": "上移跑位",
    "hot": "闷热",
    "sweaty": "出汗闷",
    "thick": "厚重",
    "not breathable": "不透气",
    "not adjustable": "肩带不可调节",
    "straps not adjustable": "肩带不可调节",
    "can't adjust": "肩带不可调节",
    "cannot adjust": "肩带不可调节",
    "straps too long": "肩带太长",
    "straps too short": "肩带太短",
    "strap digs": "肩带勒",
    "straps dig": "肩带勒",
    "strap broke": "肩带断裂",
    "straps broke": "肩带断裂",
    "strap slips": "肩带滑落",
    "straps fall": "肩带滑落",
    "color mismatch": "颜色不符",
    "not as pictured": "不如图",
    "fades": "掉色",
    "not worth": "不值",
    "overpriced": "价格偏高",
    "waste of money": "浪费钱",
}

SCENARIOS = [
    ("日常/通勤/全天穿着", ["everyday", "daily", "casual", "work", "office", "all day", "errands"]),
    ("居家/睡眠/放松", ["home", "house", "lounge", "sleep", "pajama", "around the house"]),
    ("基础内搭/外套内搭", ["under clothes", "under shirt", "under dress", "layer", "under blazer", "cardigan", "jacket"]),
    ("吊带/背心/细肩带穿搭", ["tank", "cami", "camisole", "spaghetti", "sleeveless"]),
    ("贴身/浅色/无痕内搭", ["white", "tight", "fitted", "seamless", "no lines", "show through", "thin top"]),
    ("夏季/度假/旅行", ["summer", "vacation", "travel", "beach", "hot"]),
    ("运动/瑜伽/健身", ["workout", "gym", "yoga", "exercise", "fitness"]),
    ("派对/约会/外出", ["party", "going out", "date", "club", "night out"]),
]


def safe_float(value):
    try:
        if value is None or value == "":
            return None
        return float(str(value).replace(",", "").replace("$", "").strip())
    except Exception:
        return None


def pct(n, d):
    return n / d if d else 0


def detect_asin(path):
    match = re.search(r"B0[A-Z0-9]{8}", path.name.upper())
    return match.group(0) if match else ""


def read_rows(wb, sheet):
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(v is not None and str(v).strip() for v in row):
            item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            item["__rownum"] = row_num
            rows.append(item)
    return rows


def extract_order_images(wb, asin):
    if "订单量分析" not in wb.sheetnames:
        return {}
    ws = wb["订单量分析"]
    image_dir = OUT_DIR / "images" / asin
    image_dir.mkdir(parents=True, exist_ok=True)
    by_row = {}
    for idx, image in enumerate(getattr(ws, "_images", []), start=1):
        anchor = getattr(image, "anchor", None)
        if not anchor or not hasattr(anchor, "_from"):
            continue
        row = anchor._from.row + 1
        col = anchor._from.col + 1
        if col != 1:
            continue
        suffix = Path(getattr(image, "path", "")).suffix or ".png"
        target = image_dir / f"{asin}_row{row}_{idx}{suffix}"
        try:
            target.write_bytes(image._data())
            by_row[row] = str(target)
        except Exception:
            continue
    return by_row


def first(row, names):
    for name in names:
        if row.get(name) not in (None, ""):
            return row.get(name)
    return ""


def review_text(row):
    return " ".join(str(row.get(k) or "") for k in ["标题", "标题 (翻译)", "内容", "内容(翻译)"]).replace("<br>", " ").strip()


def star(row):
    value = safe_float(row.get("星级"))
    if value and 1 <= value <= 5:
        return int(value)
    return None


def norm_size(value):
    text = str(value or "未标注").strip() or "未标注"
    mapping = {
        "X-Small": "XS", "Small": "S", "Medium": "M", "Large": "L",
        "X-Large": "XL", "XX-Large": "2XL", "XXL": "2XL", "XXX-Large": "3XL",
    }
    return mapping.get(text, text)


def parse_review_size(row):
    model = str(row.get("型号") or "")
    match = re.search(r"Size:\s*([^|]+)", model, flags=re.I)
    return norm_size(match.group(1).strip() if match else model.strip())


def original_review_text(row):
    return " ".join(str(row.get(k) or "") for k in ["标题", "内容", "Title", "Content", "Review Title", "Review Content"]).replace("<br>", " ").strip()


def extract_cup_sizes_from_text(source):
    results = []
    seen = set()

    def add(value):
        clean = re.sub(r"\s+", "", value.upper()).replace("-", "").replace("/", "")
        clean = clean.replace("DDDD", "DDDD").replace("DDD", "DDD").replace("DD", "DD")
        if clean and clean not in seen:
            seen.add(clean)
            results.append(clean)

    cup = r"(?:AA|DDDD|DDD|DD|D|A|B|C|E|F|G|H|I|J|K)"
    for band, cup_value in re.findall(rf"\b([2-5][0-9])\s*(?:-|/|\s)?\s*({cup})\b", source, flags=re.I):
        add(f"{band}{cup_value}")
    for size_value in re.findall(rf"\b(?:cup\s*size|bra\s*size|wear(?:ing)?\s*(?:a|an)?|i\s*am\s*(?:a|an)?|i'm\s*(?:a|an)?)\s*([2-5][0-9]\s*(?:-|/|\s)?\s*{cup})\b", source, flags=re.I):
        add(size_value)
    return results


def extract_review_cup_sizes(row, text):
    return extract_cup_sizes_from_text(original_review_text(row))


def extract_keyword_context_cup_sizes(row, word):
    source = original_review_text(row)
    if not source or not word:
        return []
    parts = re.split(r"(?<=[.!?。！？])\s+|[;\n\r]+", source)
    matched = [part for part in parts if word.lower() in part.lower()]
    return extract_cup_sizes_from_text(" ".join(matched))


def keyword_cell(counter):
    return "；".join(f"{word}:{count}" for word, count in counter.most_common())


def is_negated_keyword(text, word):
    patterns = [
        rf"(?:no|not|never|doesn['’]?t|does not|without|won['’]?t|will not)\s+(?:\w+\s+){{0,3}}{re.escape(word)}",
        rf"{re.escape(word)}\s+(?:\w+\s+){{0,3}}(?:no|not|never)",
    ]
    return any(re.search(pattern, text, flags=re.I) for pattern in patterns)


def keyword_occurrences(text, word):
    if " " in word:
        pattern = re.escape(word)
    else:
        pattern = rf"\b{re.escape(word)}\b"
    return [match for match in re.finditer(pattern, text, flags=re.I)]


def keyword_hit_count(text, word, sentiment=None):
    matches = keyword_occurrences(text, word)
    if sentiment == "差评":
        matches = [match for match in matches if not is_negated_keyword(text[max(0, match.start() - 40): match.end() + 40], word)]
    if sentiment == "好评" and word in {"no digging", "no marks", "no lines"}:
        return len(matches)
    return len(matches)


def keyword_context(text, word, radius=260):
    matches = keyword_occurrences(text, word)
    if not matches:
        return text
    match = matches[0]
    start = max(0, match.start() - radius)
    end = min(len(text), match.end() + radius)
    return text[start:end]


def size_cell(*counters):
    total = Counter()
    for counter in counters:
        total.update(counter)
    return "；".join(f"{size}:{count}" for size, count in total.most_common(8) if size and size != "未标注")


def merge_size_text(rows):
    total = Counter()
    for row in rows:
        source_text = row.get("提及罩杯码数", row.get("提及尺码", ""))
        for part in str(source_text or "").split("；"):
            if ":" not in part:
                continue
            size, count = part.rsplit(":", 1)
            try:
                total[size] += int(float(count))
            except Exception:
                continue
    return "；".join(f"{size}:{count}" for size, count in total.most_common(8) if size and size != "未标注")


def specific_voc_label(base_label, counter, sentiment):
    top_word = counter.most_common(1)[0][0] if counter else ""
    if sentiment == "好评" and top_word in POSITIVE_TOP_WORD_LABELS:
        return POSITIVE_TOP_WORD_LABELS[top_word]
    if sentiment == "差评" and top_word in NEGATIVE_TOP_WORD_LABELS:
        return NEGATIVE_TOP_WORD_LABELS[top_word]
    label_rules = POSITIVE_KEYWORD_LABELS if sentiment == "好评" else NEGATIVE_KEYWORD_LABELS
    words = [word for word, _ in counter.most_common()]
    for keys, label in label_rules:
        if any(key in words for key in keys):
            return label
    if sentiment == "差评":
        return negative_label(base_label)
    return base_label


def top_cell(counter, total, n=5):
    return "；".join(f"{label}:{count}({pct(count, total):.0%})" for label, count in counter.most_common(n))


def scenario_keywords(label):
    for name, words in SCENARIOS:
        if name == label:
            return "；".join(words[:8])
    return ""


def outfit(label):
    return {
        "基础内搭/外套内搭": "衬衫、开衫、西装、外套内搭",
        "吊带/背心/细肩带穿搭": "吊带、背心、细肩带上衣",
        "贴身/浅色/无痕内搭": "白色上衣、贴身T恤、浅色裙装",
        "日常/通勤/全天穿着": "通勤上衣、日常休闲装",
        "夏季/度假/旅行": "夏季短上衣、度假穿搭、旅行行李",
        "运动/瑜伽/健身": "瑜伽服、健身上衣",
        "派对/约会/外出": "约会装、派对上衣、外穿搭配",
        "居家/睡眠/放松": "居家服、睡衣、休闲装",
    }.get(label, "评论未明确搭配")


def infer_skin_tone(data):
    text = " ".join([color for color, _ in data["order_colors"].most_common(8)]).lower()
    if any(word in text for word in ["white", "ivory", "cream"]):
        return "偏好浅色/白色系，关注浅色不透与杯垫显形风险"
    if any(word in text for word in ["black", "dark"]):
        return "偏好黑色/深色基础款，重视百搭与显瘦"
    if any(word in text for word in ["nude", "beige", "tan", "brown", "skin"]):
        return "偏好肤色/裸色系，重视贴身内搭隐形效果"
    return "评论与订单未明确肤色；建议后续结合买家晒图继续判断"


def infer_body_params(data):
    sizes = "；".join(f"{size}({int(qty)})" for size, qty in data["order_sizes"].most_common(5))
    review_sizes = "；".join(f"{size}({count})" for size, count in data["review_sizes"].most_common(5))
    support = data["pos"].get("支撑承托/DD+适配", 0) + data["neg"].get("支撑承托/DD+适配", 0)
    if support:
        return f"核心订单码数：{sizes}；评论尺码：{review_sizes}；存在胸部支撑/DD+适配讨论，需覆盖B-DD+试穿。"
    return f"核心订单码数：{sizes}；评论尺码：{review_sizes}；体型信息以S/M/L订单与评论型号为主。"


def persona_text(data, age_group):
    good_total = sum(data["pos"].values())
    bad_total = sum(data["neg"].values())
    scene_total = sum(data["scenarios"].values())
    top_good = top_cell(data["pos"], good_total, 3) or "暂无集中好评"
    top_bad = top_cell(data["neg"], bad_total, 5) or "暂无集中差评"
    top_scenes = top_cell(data["scenarios"], scene_total, 5) or "评论未明显提及"
    pairings = "；".join(outfit(scene) for scene, _ in data["scenarios"].most_common(4)) or "需结合买家晒图补充"
    pref = []
    for label in ["内置罩杯/免穿内衣", "贴肤舒适/无勒痕无刺激", "遮点/浅色不透安全", "版型显瘦/胸型自然", "无痕隐形/衣服下平整"]:
        if data["pos"].get(label) or data["neg"].get(label):
            pref.append(label)
    return {
        "ASIN": data["asin"],
        "品牌": data["brand"],
        "商品标题": data["title"],
        "年龄段": age_group,
        "用户标签": "免内衣背心用户 / 舒适显身材 / 可内搭可外穿",
        "用户画像她是什么样的人": f"{age_group} 女性，想用一件带罩杯背心减少内衣选择，同时兼顾舒适、遮点、显身材和场景穿搭。",
        "核心关注的点": top_good,
        "肤色": infer_skin_tone(data),
        "体型参数": infer_body_params(data),
        "产品/设计偏好": "；".join(pref) or "评论未形成单一设计偏好",
        "核心购买Top3": top_good,
        "未满足点": top_bad,
        "使用场景": top_scenes,
        "服装搭配": pairings,
    }


def group_persona_text(group, items):
    pos = Counter()
    neg = Counter()
    scenarios = Counter()
    order_sizes = Counter()
    order_colors = Counter()
    for data in items:
        pos.update(data["pos"])
        neg.update(data["neg"])
        scenarios.update(data["scenarios"])
        order_sizes.update(data["order_sizes"])
        order_colors.update(data["order_colors"])
    good_total = sum(pos.values())
    bad_total = sum(neg.values())
    scene_total = sum(scenarios.values())
    top_good = top_cell(pos, good_total, 3) or "暂无集中好评"
    top_bad = top_cell(neg, bad_total, 5) or "暂无集中差评"
    top_scenes = top_cell(scenarios, scene_total, 5) or "评论未明显提及"
    pairings = "；".join(outfit(scene) for scene, _ in scenarios.most_common(5)) or "需结合买家晒图补充"
    skin_data = {
        "order_colors": order_colors,
        "order_sizes": order_sizes,
        "review_sizes": Counter(),
        "pos": pos,
        "neg": neg,
    }
    body = f"核心订单码数：{'；'.join(f'{s}({int(q)})' for s, q in order_sizes.most_common(6))}"
    return (
        f"1、年龄段，用户标签：{group}；免内衣背心用户 / 舒适显身材 / 可内搭可外穿。\n"
        f"用户画像她是什么样的一个人？{group} 女性，想用一件带罩杯背心减少内衣选择，同时兼顾舒适、遮点、显身材和场景穿搭。核心关注：{top_good}\n"
        f"2、肤色：{infer_skin_tone(skin_data)}\n"
        f"3、体型参数：{body}\n"
        f"4、产品/设计偏好：{top_good}\n"
        f"5、核心购买 Top3：{top_good}\n"
        f"6、未满足点：{top_bad}\n"
        f"7、使用场景：{top_scenes}\n"
        f"8、服装搭配：{pairings}"
    )


def find_files():
    wanted = set(sum(GROUPS.values(), []))
    files = {}
    for directory in SOURCE_DIRS:
        if not directory.exists():
            continue
        for path in directory.glob("*.xlsx"):
            asin = detect_asin(path)
            if asin in wanted and asin not in files:
                files[asin] = path
    return files


def analyze_asin(asin, path):
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    reviews = read_rows(wb, "评论分析")
    orders = read_rows(wb, "订单量分析")
    row_images = extract_order_images(wb, asin)
    title = ""
    brand = ""
    if orders:
        title = str(first(orders[0], ["商品标题", "Title", "标题"]) or "")
        brand = title.split()[0] if title else ""
    valid_reviews = []
    extracted_reviews = set()
    stars = Counter()
    pos = Counter()
    neg = Counter()
    pos_words = defaultdict(Counter)
    neg_words = defaultdict(Counter)
    pos_sizes = defaultdict(Counter)
    neg_sizes = defaultdict(Counter)
    evidence = []
    unmatched = []
    scenarios = Counter()
    scenario_evidence = []
    review_sizes = Counter()
    review_size_star = defaultdict(Counter)

    for idx, row in enumerate(reviews, start=2):
        s = star(row)
        text = review_text(row)
        text_l = text.lower()
        if not s or not text_l:
            continue
        valid_reviews.append(row)
        stars[s] += 1
        size = parse_review_size(row)
        cup_sizes = extract_review_cup_sizes(row, text)
        review_sizes[size] += 1
        review_size_star[size][s] += 1
        row_scenarios = []
        for scenario, words in SCENARIOS:
            if any(word in text_l for word in words):
                scenarios[scenario] += 1
                row_scenarios.append(scenario)
                if len(scenario_evidence) < 200:
                    scenario_evidence.append((scenario, s, text[:240]))
        hit = False
        for rule in TAXONOMY:
            sentiment_words = []
            for word in rule["pos"]:
                if keyword_hit_count(text_l, word, "好评"):
                    sentiment_words.append(("好评", word))
            for word in rule["neg"]:
                if keyword_hit_count(text_l, word, "差评"):
                    sentiment_words.append(("差评", word))
            sentiments_seen = set()
            for sentiment, word in sentiment_words:
                key = (idx, rule["label"], sentiment)
                if key in sentiments_seen:
                    continue
                sentiments_seen.add(key)
                hit = True
                extracted_reviews.add(idx)
                context_cup_sizes = extract_keyword_context_cup_sizes(row, word)
                hit_count = keyword_hit_count(text_l, word, sentiment)
                if not hit_count:
                    continue
                if sentiment == "好评":
                    pos[rule["label"]] += hit_count
                    pos_words[rule["label"]][word] += hit_count
                    for cup_size in context_cup_sizes:
                        pos_sizes[rule["label"]][cup_size] += hit_count
                else:
                    neg[rule["label"]] += hit_count
                    neg_words[rule["label"]][word] += hit_count
                    for cup_size in context_cup_sizes:
                        neg_sizes[rule["label"]][cup_size] += hit_count
                if len(evidence) < 600:
                    evidence.append({
                        "ASIN": asin, "品牌": brand, "VOC标签": rule["label"], "情绪": sentiment, "星级": s,
                        "评论产品码数": size,
                        "评论型号原文": str(row.get("型号") or ""),
                        "评论提及罩杯码数": "；".join(cup_sizes),
                        "命中关键词": word,
                        "命中原因": f"{sentiment}关键词命中：{word}",
                        "关键词上下文": keyword_context(text, word),
                        "原文证据片段": text,
                        "使用场景提取": "；".join(row_scenarios), "评论链接": row.get("链接") or "",
                    })
        if not hit and len(unmatched) < 300:
            unmatched.append({
                "ASIN": asin, "品牌": brand, "星级": s,
                "评论产品码数": size,
                "评论型号原文": str(row.get("型号") or ""),
                "评论提及罩杯码数": "；".join(cup_sizes),
                "评论片段": text[:260], "使用场景提取": "；".join(row_scenarios),
            })

    order_sizes = Counter()
    order_colors = Counter()
    color_images = {}
    total_orders = 0
    for row in orders:
        qty = safe_float(first(row, ["近30天订单量", "近30天销量", "订单量"])) or 0
        size = norm_size(first(row, ["Size (尺寸)", "尺码", "Size"]))
        color = str(first(row, ["Color (颜色)", "颜色", "Color"]) or "未标注").strip() or "未标注"
        image_link = str(first(row, ["图片 (数据来源于西柚洞察)", "图片", "Image", "image"]) or "").strip()
        if not image_link:
            image_link = row_images.get(row.get("__rownum"), "")
        order_sizes[size] += qty
        order_colors[color] += qty
        if image_link and color not in color_images:
            color_images[color] = image_link
        total_orders += qty

    return {
        "asin": asin, "path": path, "title": title, "brand": brand, "reviews": reviews,
        "valid_count": len(valid_reviews), "extracted_count": len(extracted_reviews), "stars": stars,
        "pos": pos, "neg": neg, "pos_words": pos_words, "neg_words": neg_words,
        "pos_sizes": pos_sizes, "neg_sizes": neg_sizes, "evidence": evidence, "unmatched": unmatched,
        "scenarios": scenarios, "scenario_evidence": scenario_evidence, "review_sizes": review_sizes,
        "review_size_star": review_size_star, "order_sizes": order_sizes, "order_colors": order_colors,
        "color_images": color_images, "total_orders": total_orders,
    }


def deep_analysis(label, good, bad):
    rule = next((item for item in TAXONOMY if item["label"] == label), None)
    mechanism = rule["mechanism"] if rule else "需结合评论和试穿继续验证。"
    if bad > good:
        return f"Confirmed VOC：{label} 的负向反馈更集中，可能影响退货、复购和外穿信心。结构逻辑：{mechanism}"
    return f"Confirmed VOC：{label} 是购买满意度驱动，应保留并在页面卖点、版型和样衣测试中强化。结构逻辑：{mechanism}"


def negative_label(label):
    return NEGATIVE_LABELS.get(label, f"{label}负向痛点")


def rice_priority(label, good, bad, total_orders):
    impact = "高" if bad and any(k in label for k in ["支撑", "尺码", "遮点", "底摆", "舒适"]) else ("中" if bad else "中低")
    reach = "高" if good + bad >= 20 or total_orders >= 1000 else ("中" if good + bad >= 6 else "低")
    confidence = "高" if good + bad >= 10 else "中"
    effort = "高" if any(k in label for k in ["支撑", "罩杯", "尺码"]) else "中"
    priority = "P1" if impact == "高" and confidence in {"高", "中"} else ("P2" if good + bad >= 5 else "P3")
    return f"Reach:{reach}；Impact:{impact}；Confidence:{confidence}；Effort:{effort}；优先级:{priority}"


def voc_rows_with_asin_summary(rows):
    sorted_rows = sorted(
        rows,
        key=lambda r: (
            str(r.get("ASIN", "")),
            r.get("好评频次", 0) + r.get("差评频次", 0),
            r.get("好评频次", 0),
        ),
        reverse=True,
    )
    output = []
    grouped = defaultdict(list)
    for row in sorted_rows:
        grouped[row.get("ASIN", "")].append(row)

    for asin in sorted(grouped.keys(), reverse=True):
        asin_rows = grouped[asin]
        output.extend(asin_rows)
        first_row = asin_rows[0]
        good_count = sum(row.get("好评频次", 0) for row in asin_rows)
        bad_count = sum(row.get("差评频次", 0) for row in asin_rows)
        good_share = sum(row.get("好评占比", 0) for row in asin_rows)
        bad_share = sum(row.get("差评占比", 0) for row in asin_rows)
        output.append({
            "ASIN": asin,
            "品牌": first_row.get("品牌", ""),
            "商品标题": first_row.get("商品标题", ""),
            "好评标签（中文）": "该ASIN汇总",
            "概括好评关键词（原文英文）": "",
            "好评频次": good_count,
            "好评占比": good_share,
            "好评评论杯码": "",
            "对应的差评关键词（中文）": "该ASIN差评汇总",
            "差评关键词（英文原文）": "",
            "差评频次": bad_count,
            "差评占比": bad_share,
            "差评评论杯码": "",
            "提及罩杯码数": merge_size_text(asin_rows),
            "深度分析（逻辑）": (
                f"该ASIN VOC汇总：好评频次{good_count}，占比{good_share:.2%}；"
                f"差评频次{bad_count}，占比{bad_share:.2%}；好评+差评占比合计{good_share + bad_share:.2%}。"
            ),
            "Kano属性": "汇总",
            "RICE/优先级": "见明细",
        })
    return output


def source_anomaly_text(data):
    issues = []
    if not data["title"]:
        issues.append("商品标题缺失")
    if data["valid_count"] == 0:
        issues.append("评论表无有效评论")
    if data["extracted_count"] == 0:
        issues.append("有效评论未命中VOC分类")
    if data["total_orders"] == 0:
        issues.append("订单量分析无订单量")
    if not data["order_sizes"]:
        issues.append("码数订单数据缺失")
    if not data["order_colors"]:
        issues.append("颜色订单数据缺失")
    if data["order_colors"] and not any(data["color_images"].values()):
        issues.append("颜色图片未提取到")
    return "正常" if not issues else "；".join(issues)


def style_sheet(ws):
    orange = PatternFill("solid", fgColor="ED7D31")
    white = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    body = Font(name="Microsoft YaHei", size=10, color="1F2933")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = orange
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        is_summary = any(cell.value == "该ASIN汇总" for cell in row)
        for cell in row:
            cell.font = body
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if is_summary:
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="1F2933")
            if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                cell.number_format = "0.00%"
    ws.auto_filter.ref = ws.dimensions
    for idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value or "")
        width = 16
        if any(k in header for k in ["标题", "关键词", "分析", "证据", "片段", "画像", "问题", "文件", "链接", "场景"]):
            width = 46
        ws.column_dimensions[get_column_letter(idx)].width = width


def resolve_image_path(value):
    if not value:
        return None
    path = Path(str(value))
    if not path.is_absolute():
        path = Path.cwd() / path
    return path if path.exists() and path.is_file() else None


def embed_representative_images(ws, headers):
    if "代表图片" not in headers:
        return
    col = headers.index("代表图片") + 1
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = 18
    for row_idx in range(2, ws.max_row + 1):
        path = resolve_image_path(ws.cell(row=row_idx, column=col).value)
        if not path:
            continue
        try:
            image = XLImage(str(path))
            image.width = 82
            image.height = 82
            ws.add_image(image, f"{col_letter}{row_idx}")
            ws.cell(row=row_idx, column=col).value = ""
            ws.row_dimensions[row_idx].height = 66
        except Exception:
            continue


def add_sheet(wb, name, rows):
    ws = wb.create_sheet(name[:31])
    headers = []
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
    else:
        ws.append(["暂无数据"])
    style_sheet(ws)
    embed_representative_images(ws, headers)


def build_workbook(data_by_asin, files):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    SOURCE_COPY.mkdir(exist_ok=True)
    for path in files.values():
        shutil.copy2(path, SOURCE_COPY / path.name)

    overview = []
    voc_rows = []
    evidence_rows = []
    unmatched_rows = []
    demand_rows = []
    asin_demand_rows = []
    size_rows = []
    color_rows = []
    color_simple_rows = []
    all_color = Counter()
    all_color_images = {}
    all_color_asins = defaultdict(Counter)
    demand_total = Counter()
    demand_good = Counter()
    demand_bad = Counter()
    scenario_all = Counter()
    scenario_by_group = defaultdict(Counter)
    scenario_detail = []
    persona_rows = []
    group_persona_source = defaultdict(list)
    translation_rows = []
    source_rows = []
    anomaly_rows = []

    for asin, data in data_by_asin.items():
        good_reviews = data["stars"][4] + data["stars"][5]
        bad_reviews = data["stars"][1] + data["stars"][2] + data["stars"][3]
        view_total = sum(data["pos"].values()) + sum(data["neg"].values())
        overview.append({
            "ASIN": asin, "品牌": data["brand"], "商品标题": data["title"], "有效评论数": data["valid_count"],
            "提取有效评论数": data["extracted_count"], "123星数量": bad_reviews, "4.5星数量": good_reviews,
            "差评占比": pct(bad_reviews, data["valid_count"]),
        })
        labels = set(data["pos"]) | set(data["neg"])
        good_total = sum(data["pos"].values())
        bad_total = sum(data["neg"].values())
        for label in sorted(labels, key=lambda x: (data["pos"].get(x, 0) + data["neg"].get(x, 0), data["pos"].get(x, 0), data["neg"].get(x, 0)), reverse=True):
            good = data["pos"].get(label, 0)
            bad = data["neg"].get(label, 0)
            demand_count = good + bad
            rule = next((item for item in TAXONOMY if item["label"] == label), {})
            good_label = specific_voc_label(label, data["pos_words"][label], "好评") if good else ""
            bad_label = specific_voc_label(label, data["neg_words"][label], "差评") if bad else ""
            voc_rows.append({
                "ASIN": asin, "品牌": data["brand"], "商品标题": data["title"],
                "好评标签（中文）": good_label,
                "概括好评关键词（原文英文）": keyword_cell(data["pos_words"][label]),
                "好评频次": good, "好评占比": pct(good, view_total),
                "好评评论杯码": keyword_cell(data["pos_sizes"][label]),
                "对应的差评关键词（中文）": bad_label,
                "差评关键词（英文原文）": keyword_cell(data["neg_words"][label]),
                "差评频次": bad, "差评占比": pct(bad, view_total),
                "差评评论杯码": keyword_cell(data["neg_sizes"][label]),
                "提及罩杯码数": size_cell(data["pos_sizes"][label], data["neg_sizes"][label]),
                "深度分析（逻辑）": deep_analysis(label, good, bad),
                "Kano属性": rule.get("kano", ""),
                "RICE/优先级": rice_priority(label, good, bad, data["total_orders"]),
            })
            demand_total[label] += good + bad
            demand_good[label] += good
            demand_bad[label] += bad
            asin_demand_rows.append({
                "ASIN": asin,
                "品牌": data["brand"],
                "商品标题": data["title"],
                "需求标签": label,
                "好评次数": good,
                "好评占比": pct(good, good_total),
                "差评次数": bad,
                "差评占比": pct(bad, bad_total),
                "汇总需求次数": demand_count,
                "汇总需求占比": pct(demand_count, view_total),
                "提及罩杯码数": size_cell(data["pos_sizes"][label], data["neg_sizes"][label]),
                "Kano属性": rule.get("kano", ""),
                "开发判断": deep_analysis(label, good, bad),
                "RICE/优先级": rice_priority(label, good, bad, data["total_orders"]),
            })
        evidence_rows.extend(data["evidence"])
        unmatched_rows.extend(data["unmatched"])
        demand_rows.append({
            "ASIN": asin, "品牌": data["brand"], "有效评论数": data["valid_count"],
            "45星评论数": good_reviews, "45星评论占比": pct(good_reviews, data["valid_count"]),
            "123星评论数": bad_reviews, "123星评论占比": pct(bad_reviews, data["valid_count"]),
            "好评观点总频次": good_total, "Top5好评需求层级/占比": top_cell(data["pos"], good_total),
            "差评观点总频次": bad_total, "Top5核心差评点/占比": top_cell(data["neg"], bad_total),
        })
        for size, qty in data["order_sizes"].most_common():
            review_count = data["review_sizes"].get(size, 0)
            bad_size = data["review_size_star"][size][1] + data["review_size_star"][size][2] + data["review_size_star"][size][3]
            good_size = data["review_size_star"][size][4] + data["review_size_star"][size][5]
            size_rows.append({
                "ASIN": asin, "品牌": data["brand"], "有效评论数": data["valid_count"],
                "45星评论数": good_reviews, "45星评论占比": pct(good_reviews, data["valid_count"]),
                "123星评论数": bad_reviews, "123星评论占比": pct(bad_reviews, data["valid_count"]),
                "好评观点总频次": good_total, "差评观点总频次": bad_total,
                "Top好评需求": data["pos"].most_common(1)[0][0] if data["pos"] else "",
                "Top核心差评点": data["neg"].most_common(1)[0][0] if data["neg"] else "",
                "码数": size, "完整尺码": "", "订单量": int(round(qty)), "订单量占比": pct(qty, data["total_orders"]),
                "该码数评论数": review_count, "123星数量": bad_size, "123星占比": pct(bad_size, data["valid_count"]),
                "45星数量": good_size, "45星占比": pct(good_size, data["valid_count"]),
                "评论正文数字码数提及": "", "数字码数映射依据": "订单Size字段 + 评论型号Size字段；完整杯码需尺码图补充",
            })
        for color, qty in data["order_colors"].most_common():
            image_link = data["color_images"].get(color, "")
            color_rows.append({
                "代表图片": image_link, "ASIN": asin, "品牌/商品标题": f"{data['brand']} / {data['title']}",
                "颜色": color, "订单量": int(round(qty)), "订单量占比": pct(qty, data["total_orders"]),
                "评论数/观点数（如有）": data["valid_count"], "近30天订单量（如有）": int(round(qty)),
                "近30天下单占比（如有）": pct(qty, data["total_orders"]), "图片文件": image_link, "源文件": str(data["path"]),
            })
            color_simple_rows.append({
                "代表图片": image_link, "ASIN": asin, "商品标题": data["title"], "颜色名称": color,
                "订单量": int(round(qty)), "订单量占比": pct(qty, data["total_orders"]), "图片链接": image_link,
            })
            all_color[color] += qty
            all_color_asins[color][asin] += qty
            if image_link and color not in all_color_images:
                all_color_images[color] = image_link
        age_group = next((g for g, vals in GROUPS.items() if asin in vals), "")
        scenario_total = sum(data["scenarios"].values())
        scenario_all.update(data["scenarios"])
        scenario_by_group[age_group].update(data["scenarios"])
        for scenario, count in data["scenarios"].most_common():
            scenario_detail.append({
                "年龄段": age_group, "ASIN": asin, "品牌": data["brand"], "使用场景标签": scenario, "代表关键词": scenario_keywords(scenario),
                "服装搭配": outfit(scenario), "提及次数": count, "占比": pct(count, scenario_total),
            })
        group_persona_source[age_group].append(data)
        source_rows.append({"ASIN": asin, "品牌": data["brand"], "源文件": str(data["path"]), "状态": "已读取"})
        anomaly_rows.append({
            "ASIN": asin,
            "品牌": data["brand"],
            "源文件": str(data["path"]),
            "有效评论数": data["valid_count"],
            "提取有效评论数": data["extracted_count"],
            "订单量合计": int(round(data["total_orders"])),
            "颜色数": len(data["order_colors"]),
            "码数数": len(data["order_sizes"]),
            "颜色图片数": sum(1 for value in data["color_images"].values() if value),
            "异常状态": source_anomaly_text(data),
        })

    all_color_total = sum(all_color.values())
    all_color_rows = [
        {"代表图片": all_color_images.get(color, ""), "颜色": color, "近30天总订单量": int(round(qty)), "近30天销量占比": pct(qty, all_color_total),
         "期间总订单量": int(round(qty)), "期间销量占比": pct(qty, all_color_total), "涉及ASIN数": "",
         "涉及源文件数": "", "主要贡献ASIN(按近30天)": "；".join(asin for asin, _ in all_color_asins[color].most_common(5)),
         "主要品牌": "", "图片文件": all_color_images.get(color, "")}
        for color, qty in all_color.most_common()
    ]
    demand_grand = sum(demand_total.values())
    good_grand = sum(demand_good.values())
    bad_grand = sum(demand_bad.values())
    demand_total_rows = [
        {"需求标签": label, "总提及次数": count, "总占比": pct(count, demand_grand),
         "好评提及次数": demand_good[label], "好评占比": pct(demand_good[label], good_grand),
         "差评提及次数": demand_bad[label], "差评占比": pct(demand_bad[label], bad_grand),
         "Kano属性": next((r["kano"] for r in TAXONOMY if r["label"] == label), ""),
         "开发判断": deep_analysis(label, demand_good[label], demand_bad[label]),
         "RICE/优先级": rice_priority(label, demand_good[label], demand_bad[label], 0)}
        for label, count in demand_total.most_common()
    ]
    scenario_grand = sum(scenario_all.values())
    scenario_total_rows = []
    for group, counter in scenario_by_group.items():
        group_total = sum(counter.values())
        for name, count in counter.most_common():
            scenario_total_rows.append({
                "年龄段": group, "使用场景标签": name, "代表关键词": scenario_keywords(name),
                "服装搭配": outfit(name), "提及次数": count, "占比": pct(count, group_total),
            })
    scenario_total_rows.extend([
        {"年龄段": "全部ASIN汇总", "使用场景标签": name, "代表关键词": scenario_keywords(name), "服装搭配": outfit(name), "提及次数": count, "占比": pct(count, scenario_grand)}
        for name, count in scenario_all.most_common()
    ])
    persona_rows = []
    for group, items in group_persona_source.items():
        for data in items:
            row = persona_text(data, group)
            order_size_total = sum(data["order_sizes"].values())
            review_size_total = sum(data["review_sizes"].values())
            order_size_top = "；".join(
                f"{size}:{int(qty)}({pct(qty, order_size_total):.0%})"
                for size, qty in data["order_sizes"].most_common(8)
            )
            review_size_top = "；".join(
                f"{size}:{count}({pct(count, review_size_total):.0%})"
                for size, count in data["review_sizes"].most_common(8)
                if size and size != "未标注"
            )
            row["订单尺码Top"] = order_size_top
            row["评论提及尺码/罩杯"] = review_size_top
            row["围度/尺码判断"] = infer_body_params(data)
            row["判断依据"] = (
                f"年龄段来自该ASIN分组；订单尺码Top={order_size_top or '无订单尺码'}；"
                f"评论提及尺码/罩杯={review_size_top or '评论未明确尺码'}；"
                f"核心好评={top_cell(data['pos'], sum(data['pos'].values()), 3) or '无'}；"
                f"核心差评={top_cell(data['neg'], sum(data['neg'].values()), 3) or '无'}。"
            )
            persona_rows.append(row)
    for rule in TAXONOMY:
        translation_rows.append({
            "中文标签": rule["label"], "正向英文关键词": "；".join(rule["pos"]), "负向英文关键词": "；".join(rule["neg"]),
            "产品机制": rule["mechanism"], "Kano属性": rule["kano"],
        })

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "01_竞对好评差评总览", overview)
    add_sheet(
        wb,
        "02_VOC-ASIN好评差评分析",
        voc_rows_with_asin_summary(voc_rows),
    )
    add_sheet(wb, "03_评论证据样例", evidence_rows)
    add_sheet(wb, "04_未归类观点检查", unmatched_rows)
    add_sheet(wb, "05_各ASIN需求层级摘要", demand_rows)
    add_sheet(wb, "06_ASIN码数订单评论", size_rows)
    add_sheet(wb, "07_ASIN颜色订单占比", color_rows)
    add_sheet(wb, "08_ASIN颜色统计表", color_simple_rows)
    add_sheet(wb, "09_所有ASIN颜色汇总", all_color_rows)
    add_sheet(wb, "10_需求汇总统计", demand_total_rows)
    add_sheet(wb, "10A_各ASIN需求汇总统计", asin_demand_rows)
    add_sheet(wb, "10B_数据源异常检查", anomaly_rows)
    add_sheet(wb, "11_使用场景汇总", scenario_total_rows)
    add_sheet(wb, "12_各ASIN使用场景明细", scenario_detail)
    add_sheet(wb, "13_用户画像问题建议", persona_rows)
    add_sheet(wb, "标签翻译表", translation_rows)
    add_sheet(wb, "来源文件", source_rows)
    wb.save(OUT_FILE)
    return OUT_FILE


def main():
    files = find_files()
    missing = [asin for vals in GROUPS.values() for asin in vals if asin not in files]
    data = {asin: analyze_asin(asin, path) for asin, path in files.items() if asin in sum(GROUPS.values(), [])}
    output = build_workbook(data, files)
    check = openpyxl.load_workbook(output, read_only=True)
    print(output.resolve())
    print(check.sheetnames)
    print("missing", missing)


if __name__ == "__main__":
    main()
