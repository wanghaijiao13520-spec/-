from __future__ import annotations

from datetime import datetime
from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path
import re
import shutil
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from report_engine import JOBS_DIR
    from llm_client import chat_completions, enabled as llm_enabled, sanitize_status
except ModuleNotFoundError:
    from webapp.report_engine import JOBS_DIR
    from webapp.llm_client import chat_completions, enabled as llm_enabled, sanitize_status


ROOT = Path(__file__).resolve().parents[1]
TOOLS_DIR = ROOT / "tools"

SCRIPT_MAP = {
    "tank": {
        "path": TOOLS_DIR / "build_voc_product_expert_tank.py",
        "name": "背心VOC-PRODUCT",
        "file_prefix": "VOC_PRODUCT_背心产品专家分析",
    },
    "sexy_sheer": {
        "path": TOOLS_DIR / "build_voc_product_expert_sexy_sheer.py",
        "name": "性感透视款VOC-PRODUCT",
        "file_prefix": "VOC_PRODUCT_性感透视款产品专家分析",
    },
    "plus_size_thin_cup": {
        "path": TOOLS_DIR / "build_voc_product_expert_plus_size_thin_cup.py",
        "name": "大码薄杯款VOC-PRODUCT",
        "file_prefix": "VOC_PRODUCT_大码薄杯款产品专家分析",
    },
}


def detect_asin_from_file(path: Path) -> str:
    match = re.search(r"B0[A-Z0-9]{8}", path.name.upper())
    if match:
        return match.group(0)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets[:4]:
            for row in ws.iter_rows(max_row=8, values_only=True):
                for value in row:
                    match = re.search(r"B0[A-Z0-9]{8}", str(value or "").upper())
                    if match:
                        return match.group(0)
    except Exception:
        return ""
    return ""


def unique_asins(files: list[Path]) -> list[str]:
    seen = []
    for path in files:
        asin = detect_asin_from_file(path)
        if asin and asin not in seen:
            seen.append(asin)
    return seen


def load_builder(report_type: str):
    config = SCRIPT_MAP.get(report_type) or SCRIPT_MAP["tank"]
    spec = spec_from_file_location(f"voc_builder_{report_type}_{datetime.now().strftime('%H%M%S%f')}", config["path"])
    if not spec or not spec.loader:
        raise RuntimeError("无法加载VOC-PRODUCT生成逻辑")
    module = module_from_spec(spec)
    spec.loader.exec_module(module)
    return module, config


def read_sheet_preview(path: Path, sheet_name: str, limit: int = 60) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for row in rows[1 : limit + 1]:
        if not any(value not in (None, "") for value in row):
            continue
        result.append({headers[i] or f"字段{i + 1}": row[i] for i in range(min(len(headers), len(row)))})
    return result


def style_sheet(ws):
    orange = PatternFill("solid", fgColor="ED7D31")
    white = Font(color="FFFFFF", bold=True, name="Microsoft YaHei", size=10)
    body = Font(name="Microsoft YaHei", size=10, color="1F2933")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = orange
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.auto_filter.ref = ws.dimensions
    for idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value or "")
        ws.column_dimensions[get_column_letter(idx)].width = 42 if any(k in header for k in ["画像", "依据", "分析", "证据", "场景", "标题"]) else 20


def export_sheet_workbooks(source: Path, job_dir: Path) -> dict[str, str]:
    source_wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    downloads = {}
    for sheet_name in source_wb.sheetnames:
        rows = list(source_wb[sheet_name].iter_rows(values_only=True))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        if rows:
            for row in rows:
                ws.append(list(row))
        else:
            ws.append(["暂无数据"])
        style_sheet(ws)
        filename = f"{sheet_name[:28]}.xlsx"
        target = job_dir / filename
        wb.save(target)
        downloads[sheet_name] = f"/download/{job_dir.name}/{filename}"
    return downloads


def compact_rows(rows: list[dict[str, Any]], limit: int = 12) -> list[dict[str, Any]]:
    compact = []
    keep_words = ["ASIN", "品牌", "标题", "标签", "关键词", "频次", "占比", "需求", "痛点", "画像", "尺码", "罩杯", "订单", "场景", "判断"]
    for row in rows[:limit]:
        item = {}
        for key, value in row.items():
            if any(word in str(key) for word in keep_words):
                item[str(key)] = str(value or "")[:260]
        compact.append(item)
    return compact


def build_llm_prompt(module_data: dict[str, list[dict[str, Any]]], asins: list[str], metadata: dict[str, str]) -> list[dict[str, str]]:
    payload = {
        "产品分类": metadata.get("productCategory", ""),
        "款号": metadata.get("styleNumber", ""),
        "ASIN列表": asins,
        "VOC分析": compact_rows(module_data.get("02_VOC-ASIN好评差评分析", []), 40),
        "各ASIN需求汇总": compact_rows(module_data.get("10A_各ASIN需求汇总统计", []), 40),
        "用户画像": compact_rows(module_data.get("13_用户画像问题建议", []), 20),
        "使用场景": compact_rows(module_data.get("12_各ASIN使用场景明细", []), 25),
    }
    system = (
        "你是内衣/服装产品VOC分析专家。只能基于用户提供的结构化统计和评论证据做分析。"
        "不得修改频次、占比、订单量、尺码统计；不得编造评论没有提及的尺码/罩杯；"
        "不得输出与产品结构冲突的建议。输出必须是JSON。"
    )
    user = (
        "请基于以下VOC-PRODUCT结构化数据，输出JSON。字段必须包含："
        "directions数组3条；asin_insights数组，每项包含asin、ai_persona、unmet_needs、development_advice、evidence_basis；"
        "global_risks数组；model_note字符串。"
        "请用中文，结论要真实代表用户需求。\n\n"
        + json_dumps(payload)
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


def json_dumps(value: Any) -> str:
    import json

    return json.dumps(value, ensure_ascii=False, indent=2, default=str)


def append_llm_to_workbook(output: Path, llm_result: dict[str, Any]) -> None:
    wb = openpyxl.load_workbook(output)
    if "AI_模型增强分析" in wb.sheetnames:
        del wb["AI_模型增强分析"]
    ws = wb.create_sheet("AI_模型增强分析", 0)
    status = sanitize_status(llm_result)
    ws.append(["项目", "内容"])
    for key, label in [
        ("status", "调用状态"),
        ("message", "状态说明"),
        ("model", "模型名称"),
        ("baseUrl", "API Base URL"),
        ("elapsedMs", "耗时ms"),
    ]:
        ws.append([label, status.get(key, "")])

    data = llm_result.get("data") or {}
    directions = data.get("directions") or []
    ws.append(["开发方向123", "\n".join(str(item) for item in directions)])
    risks = data.get("global_risks") or []
    ws.append(["全局风险", "\n".join(str(item) for item in risks)])
    ws.append(["模型备注", data.get("model_note", "")])
    style_sheet(ws)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 90

    insights = data.get("asin_insights") or []
    by_asin = {str(item.get("asin", "")).upper(): item for item in insights if isinstance(item, dict)}
    persona_name = "13_用户画像问题建议"
    if persona_name in wb.sheetnames and by_asin:
        persona = wb[persona_name]
        headers = [str(cell.value or "") for cell in persona[1]]
        asin_col = headers.index("ASIN") + 1 if "ASIN" in headers else 1
        start_col = persona.max_column + 1
        added_headers = ["AI用户画像", "AI未满足需求", "AI开发建议", "AI判断依据"]
        for offset, header in enumerate(added_headers):
            persona.cell(1, start_col + offset).value = header
        for row_idx in range(2, persona.max_row + 1):
            asin = str(persona.cell(row_idx, asin_col).value or "").upper()
            insight = by_asin.get(asin)
            if not insight:
                continue
            values = [
                insight.get("ai_persona", ""),
                insight.get("unmet_needs", ""),
                insight.get("development_advice", ""),
                insight.get("evidence_basis", ""),
            ]
            for offset, value in enumerate(values):
                persona.cell(row_idx, start_col + offset).value = value
        style_sheet(persona)
    wb.save(output)


def build_voc_product_report(files: list[Path], metadata: dict[str, str] | None = None, llm_config: dict[str, str] | None = None) -> dict[str, Any]:
    metadata = metadata or {}
    report_type = metadata.get("reportType") or "tank"
    module, config = load_builder(report_type)

    job_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    job_dir = JOBS_DIR / job_id
    input_dir = job_dir / "source_inputs"
    input_dir.mkdir(parents=True, exist_ok=True)

    uploaded = []
    for source in files:
        target = input_dir / source.name
        shutil.copy2(source, target)
        uploaded.append(target)

    asins = unique_asins(uploaded)
    group_name = metadata.get("productCategory") or config["name"]
    if not asins:
        raise RuntimeError("未识别到ASIN，请确认文件名或表格内容包含ASIN")

    module.GROUPS = {group_name: asins}
    module.SOURCE_DIRS = [input_dir]
    module.OUT_DIR = job_dir
    module.SOURCE_COPY = job_dir / "source_inputs_copy"
    module.OUT_FILE = job_dir / f"{config['file_prefix']}_{datetime.now().strftime('%H%M%S')}.xlsx"

    files_by_asin = module.find_files()
    data = {asin: module.analyze_asin(asin, path) for asin, path in files_by_asin.items()}
    output = module.build_workbook(data, files_by_asin)

    wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
    modules = wb.sheetnames
    module_data = {name: read_sheet_preview(output, name) for name in modules}

    if llm_enabled(llm_config):
        llm_result = chat_completions(llm_config or {}, build_llm_prompt(module_data, asins, metadata), timeout=90, retries=1)
        append_llm_to_workbook(output, llm_result)
    else:
        llm_result = {
            "enabled": False,
            "status": "disabled",
            "message": "未启用大模型增强，使用本地规则分析",
            "model": "",
            "baseUrl": "",
            "elapsedMs": 0,
        }

    wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
    modules = wb.sheetnames
    module_data = {name: read_sheet_preview(output, name) for name in modules}
    module_downloads = export_sheet_workbooks(output, job_dir)

    records = module_data.get("14_竞对横向数据对比表", [])
    if not records:
        records = module_data.get("13_用户画像问题建议", [])
    directions = [
        "方向1：基于VOC高频好评保留核心卖点，优先强化用户已经认可的舒适、支撑、外观或场景价值。",
        "方向2：基于差评Top痛点做结构优化，重点处理尺码、罩杯、支撑、露点/透视、闷热或做工问题。",
        "方向3：按每个ASIN的用户画像拆分开发路径，结合订单尺码Top和评论提及罩杯/尺码做样衣验证。",
    ]
    llm_data = llm_result.get("data") or {}
    if isinstance(llm_data.get("directions"), list) and llm_data["directions"]:
        directions = [str(item) for item in llm_data["directions"][:3]]

    result = {
        "jobId": job_id,
        "downloadUrl": f"/download/{job_id}/{Path(output).name}",
        "asinCount": len(data),
        "records": records,
        "directions": directions,
        "modules": modules,
        "moduleData": module_data,
        "moduleDownloads": module_downloads,
        "reportType": report_type,
        "asins": asins,
        "llm": sanitize_status(llm_result),
    }
    return result
