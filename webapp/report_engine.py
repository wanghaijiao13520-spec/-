from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import json
import re
import shutil
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
JOBS_DIR = ROOT / "outputs" / "web_report_jobs"

FIELDS = [
    "开发路径", "产品图片", "买家晒图", "产品特点", "产品图片", "竞对 ASIN", "品牌/角色", "日均销量", "小类排名", "上架时间", "结论",
    "价格($)", "结论", "评分", "结论", "评论数", "竞对人群定位", "订单", "结论1：用户画像确认", "VOC分析",
    "Top3好评需求(VOC占比)", "Top5差评痛点(VOC占比)", "TOP差评点占比", "结论2痛点定位", "难度系数", "主打词",
    "使用场景", "结论5", "我方定位", "优先级"
]

VOC_RULES = [
    ("贴肤舒适/柔软亲肤", "positive", ["comfortable", "comfy", "soft", "buttery", "smooth", "second skin", "stretchy", "skin-friendly"]),
    ("内置罩杯/免穿内衣", "positive", ["built in bra", "built-in bra", "shelf bra", "no bra", "bra-free", "padded bra", "built in cups"]),
    ("支撑/承托/大胸适配", "positive", ["supportive", "support", "holds", "lift", "held everything", "large chest", "big chest", "ddd"]),
    ("遮点/不透/浅色安全", "positive", ["not see through", "opaque", "not sheer", "double lined", "coverage", "thick enough"]),
    ("版型显瘦/胸型自然", "positive", ["flattering", "fits well", "nice shape", "cute", "sexy", "snatched", "slimming", "curves"]),
    ("尺码合身/尺码表准确", "positive", ["true to size", "perfect fit", "fits perfectly", "size was perfect", "tts"]),
    ("面料质量/做工耐用", "positive", ["quality", "well made", "nice material", "washes well", "durable", "thick fabric"]),
    ("无痕/内搭平整", "positive", ["seamless", "smooth under", "under clothes", "layer", "under shirt", "under blazer"]),
    ("透气/夏季轻薄", "positive", ["breathable", "lightweight", "cool", "summer", "not hot"]),
    ("肩带/可转换穿法", "positive", ["adjustable straps", "removable straps", "extra straps", "strapless", "convertible"]),
    ("价格/性价比", "positive", ["worth it", "great price", "affordable", "value", "buy again"]),
    ("外观颜色/时尚感", "positive", ["color", "colors", "pretty", "beautiful", "cute", "ballet core", "stylish", "elegant"]),
    ("勒感/刺痒/不舒服", "negative", ["uncomfortable", "itchy", "scratchy", "dig", "digs", "chafing", "tight", "suffocate"]),
    ("罩杯结构不稳定/不够好用", "negative", ["bra design", "cup", "cups", "padding", "pads", "pad moves", "pads move"]),
    ("支撑不足/大胸不适配", "negative", ["no support", "not supportive", "lacks support", "not enough support", "not for large", "falls"]),
    ("薄透/白色露杯/罩杯显形", "negative", ["see through", "sheer", "thin", "pad through", "pads visible", "visible", "white tank"]),
    ("胸型怪/压胸/外扩/副乳", "negative", ["weird fit", "awkward", "unflattering", "flatten", "spillage", "side boob", "armpit", "boobs"]),
    ("尺码偏小/偏大/尺码表不准", "negative", ["too small", "too big", "runs small", "runs big", "size up", "not true to size", "size chart"]),
    ("质量弱/QC/洗后问题", "negative", ["poor quality", "cheap", "seams", "stitch", "broke", "wash", "fraying", "threads"]),
    ("衣服下显线/不平整", "negative", ["lines", "indentation", "bulge", "shows under", "not smooth"]),
    ("闷热/厚重/不透气", "negative", ["hot", "sweaty", "heavy", "too thick", "not breathable"]),
    ("肩带硌/断/不稳", "negative", ["straps broke", "straps dig", "strap attachment"]),
    ("底摆卷边/上移/堆积", "negative", ["rolls", "roll up", "rolls up", "curls", "bunches", "rides up"]),
    ("不值/价格偏高", "negative", ["not worth", "overpriced", "waste of money"]),
    ("颜色不符/外观不如图", "negative", ["wrong color", "color off", "not as pictured", "different color", "ugly"]),
]

SCENARIO_RULES = [
    ("日常/通勤/全天穿着", ["everyday", "daily", "casual", "work", "office", "all day", "errands"]),
    ("居家/睡眠/放松", ["home", "house", "lounge", "sleep", "pajama", "around the house"]),
    ("基础内搭/外套内搭", ["under clothes", "under shirt", "under dress", "layer", "under blazer", "cardigan", "jacket"]),
    ("吊带/背心/细肩带穿搭", ["tank", "cami", "camisole", "spaghetti", "sleeveless"]),
    ("贴身/浅色/无痕内搭", ["white", "tight", "fitted", "seamless", "no lines", "show through", "thin top"]),
    ("夏季/度假/旅行", ["summer", "vacation", "travel", "beach", "hot"]),
    ("运动/瑜伽/健身", ["workout", "gym", "yoga", "exercise", "fitness"]),
    ("派对/约会/外出", ["party", "going out", "date", "club", "night out"]),
]

EXPERT_RULES = {
    "贴肤舒适/柔软亲肤": {
        "jtbd": "用户想全天穿着不被勒、不刺痒、不需要频繁调整。",
        "barrier": "面料手感、边口压力、内层/胸垫接触不稳定会直接变成退货和差评。",
        "mechanism": "面料纤维、弹性回复、车缝边口和内置杯接触面共同决定贴肤体验。",
        "opportunity": "优先做亲肤高弹面料、低摩擦边口和杯位无压迫试穿。",
        "kano": "基本型",
        "effort": "中",
    },
    "内置罩杯/免穿内衣": {
        "jtbd": "用户想省掉内衣选择，一件背心直接完成外穿/内搭。",
        "barrier": "杯位偏高、杯垫移位、罩杯显形会破坏免穿内衣的核心价值。",
        "mechanism": "下围弹力带、杯型固定方式、杯色和外层面料决定免内衣体验。",
        "opportunity": "做固定杯/稳定杯结构，白色杯色匹配，杯边过渡要弱化。",
        "kano": "期望型",
        "effort": "中高",
    },
    "支撑/承托/大胸适配": {
        "jtbd": "用户需要在日常活动中被托住，尤其是B-DD+用户不能下坠或晃动。",
        "barrier": "支撑不足会导致无法外穿、胸型不安全、公众场景尴尬。",
        "mechanism": "支撑来自下围、杯宽、侧翼高度、面料模量和尺码映射。",
        "opportunity": "核心码S/M/L做不同胸围试穿，强化下围和侧翼包裹。",
        "kano": "期望型",
        "effort": "高",
    },
    "遮点/不透/浅色安全": {
        "jtbd": "用户想穿白色/浅色也不露杯、不透点、不尴尬。",
        "barrier": "浅色露杯、薄透、杯边印会直接让产品只能做内搭，削弱外穿价值。",
        "mechanism": "面料克重、双层结构、杯色、杯边厚薄和领口张力共同影响透视。",
        "opportunity": "白色必须做遮点测试，双层面料和杯色匹配是页面主卖点。",
        "kano": "基本型",
        "effort": "中",
    },
    "版型显瘦/胸型自然": {
        "jtbd": "用户想要显腰、胸型自然、外穿好看，同时不压胸不副乳。",
        "barrier": "胸型怪、压胸、副乳会让用户觉得不像图片，不愿外穿。",
        "mechanism": "领型、杯宽、侧翼高度、腰腹贴合和肩带位置决定视觉比例。",
        "opportunity": "按领型做版型差异：方领显肩颈，V领修饰胸线，圆领做通勤安全。",
        "kano": "期望型",
        "effort": "中高",
    },
    "尺码合身/尺码表准确": {
        "jtbd": "用户希望按平时尺码下单就能合身，不靠退换货试错。",
        "barrier": "尺码偏差会连带造成勒感、支撑不足、杯位错位和卷边。",
        "mechanism": "S/M/L到胸围、下胸围、杯围和衣长的映射决定体验稳定性。",
        "opportunity": "尺码表要用订单核心码和评论尺码校正，页面提示是否需size up。",
        "kano": "基本型",
        "effort": "中",
    },
    "面料质量/做工耐用": {
        "jtbd": "用户希望洗后不变形、不脱线、杯垫不乱。",
        "barrier": "做工差会降低评分，尤其在带罩杯结构里会放大质量风险。",
        "mechanism": "缝线、杯垫固定、面料回弹和洗后稳定性决定复购。",
        "opportunity": "强化杯垫固定、洗后测试和边口工艺，减少低星质量差评。",
        "kano": "基本型",
        "effort": "中",
    },
    "底摆卷边/上移/堆积": {
        "jtbd": "用户想坐下、抬手、走动时背心仍服帖不卷边。",
        "barrier": "卷边会暴露腰腹、影响外穿，也会让用户频繁整理。",
        "mechanism": "衣长、底摆弹力、腰腹版型、面料克重和身体曲线适配共同影响卷边。",
        "opportunity": "增加底摆稳定、衣长和腰腹包覆试穿，特别验证S/M/L核心码。",
        "kano": "基本型",
        "effort": "中",
    },
}


def safe_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(str(value).replace("$", "").replace(",", "").strip())
    except Exception:
        return None


def pct(n: float, d: float) -> float:
    return n / d if d else 0



def keyword_occurrences(text: str, word: str) -> list[re.Match[str]]:
    pattern = r"(?<![a-z0-9])" + re.escape(word.lower()) + r"(?![a-z0-9])"
    return list(re.finditer(pattern, text.lower()))


def is_negated_keyword(text: str, match: re.Match[str]) -> bool:
    start = max(0, match.start() - 42)
    end = min(len(text), match.end() + 28)
    ctx = text[start:end].lower().replace("?", "'").replace("`", "'")
    negators = [
        "no ", "not ", "never ", "without ", "doesn't ", "does not ", "didn't ", "did not ",
        "isn't ", "is not ", "aren't ", "are not ", "won't ", "will not ", "don't ", "do not ",
        "doesnt ", "didnt ", "isnt ", "arent ", "wont ", "dont ", "can't ", "cannot ", "cant ",
    ]
    return any(neg in ctx for neg in negators)


def keyword_hit_count(text: str, word: str, sentiment: str | None = None) -> int:
    matches = keyword_occurrences(text, word)
    if sentiment == "negative":
        matches = [m for m in matches if not is_negated_keyword(text, m)]
    return len(matches)


def keyword_context(text: str, word: str, radius: int = 260) -> str:
    matches = keyword_occurrences(text, word)
    if not matches:
        return ""
    m = matches[0]
    start = max(0, m.start() - radius)
    end = min(len(text), m.end() + radius)
    return text[start:end].strip()


def matched_keyword_counts(text: str, words: list[str], sentiment: str | None = None) -> Counter:
    counts = Counter()
    for word in words:
        count = keyword_hit_count(text, word, sentiment)
        if count:
            counts[word] += count
    return counts


def fmt_num(value: float | None) -> str:
    if value is None:
        return ""
    return f"{int(round(value)):,}"


def read_sheet_dicts(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return rows


def first_present(row: dict[str, Any], names: list[str]) -> Any:
    for name in names:
        if name in row and row.get(name) not in (None, ""):
            return row.get(name)
    lower_names = [name.lower() for name in names]
    for key, value in row.items():
        key_l = str(key).lower()
        if value not in (None, "") and any(name in key_l for name in lower_names):
            return value
    return ""


def asin_from_text(value: Any) -> str:
    match = re.search(r"B0[A-Z0-9]{8}", str(value or "").upper())
    return match.group(0) if match else ""


def product_amz_url(asin: str, market: dict[str, Any], orders: list[dict[str, Any]]) -> str:
    keys = ["产品AMZ链接", "产品azm链接", "产品amz链接", "AMZ链接", "Amazon链接", "商品详情页链接", "商品链接", "链接", "URL"]
    value = first_present(market, keys)
    if not value and orders:
        value = first_present(orders[0], keys)
    if value:
        return str(value).strip()
    return f"https://www.amazon.com/dp/{asin}"


def product_image_source(wb, asin: str, amz_url: str) -> str:
    if "产品细节图" in wb.sheetnames:
        ws = wb["产品细节图"]
        links = []
        for row in ws.iter_rows(min_row=2, max_row=3, values_only=True):
            for value in row[1:6]:
                text = str(value or "").strip()
                if text and text.lower().startswith(("http://", "https://")):
                    links.append(text)
            if links:
                break
        if links:
            return "\n".join(links[:3])
    return f"AMZ前台：{amz_url}"


def review_image_evidence(rows: list[dict[str, Any]], fallback_url: str, limit: int = 3) -> str:
    snippets = []
    for row in rows:
        image_count = safe_float(row.get("图片数量")) or 0
        if image_count <= 0:
            continue
        title = str(row.get("标题 (翻译)") or row.get("标题") or "").replace("\n", " ").strip()
        content = str(row.get("内容(翻译)") or row.get("内容") or "").replace("\n", " ").strip()
        link = str(row.get("链接") or fallback_url or "").strip()
        star = str(row.get("星级") or "").strip()
        snippet = title or content[:80] or "带图评论"
        snippets.append(f"AMZ买家评论晒图{int(image_count)}图｜{star}星｜{snippet[:70]}｜{link}")
        if len(snippets) >= limit:
            break
    if snippets:
        return "\n".join(snippets)
    return f"未在导入评论表识别到带图评论；前台核验：{fallback_url}"


def detect_asin(path: Path, wb=None) -> str:
    match = re.search(r"B0[A-Z0-9]{8}", path.name.upper())
    if match:
        return match.group(0)
    if wb and "产品细节图" in wb.sheetnames:
        value = wb["产品细节图"]["A2"].value
        if value:
            return str(value).strip()
    return path.stem[:20]


def find_market_file(files: list[Path]) -> Path | None:
    for path in files:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if "tank-top" in wb.sheetnames:
                return path
        except Exception:
            continue
    return None


def market_rows(market_file: Path | None, asins: set[str]) -> dict[str, dict[str, Any]]:
    if not market_file:
        return {}
    wb = openpyxl.load_workbook(market_file, data_only=True)
    if "tank-top" not in wb.sheetnames:
        return {}
    ws = wb["tank-top"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        asin = str(item.get("ASIN") or "").strip().upper()
        if not asin:
            asin = asin_from_text(first_present(item, ["产品AMZ链接", "产品azm链接", "产品amz链接", "AMZ链接", "Amazon链接", "商品详情页链接", "商品链接", "链接", "URL"]))
        if asin in asins:
            rows[asin] = item
    return rows


def classify_reviews(rows: list[dict[str, Any]]) -> tuple[Counter, Counter, Counter, list[dict[str, Any]]]:
    pos = Counter()
    neg = Counter()
    scenarios = Counter()
    evidence = []
    for idx, row in enumerate(rows, start=2):
        text = " ".join(str(row.get(k) or "") for k in ["\u6807\u9898", "\u6807\u9898 (\u7ffb\u8bd1)", "\u5185\u5bb9", "\u5185\u5bb9(\u7ffb\u8bd1)"]).replace("<br>", " ")
        text_l = text.lower()
        if not text_l.strip():
            continue
        for label, sentiment, words in VOC_RULES:
            word_counts = matched_keyword_counts(text_l, words, sentiment)
            if not word_counts:
                continue
            (pos if sentiment == "positive" else neg)[label] += sum(word_counts.values())
            evidence.append({
                "\u8bc4\u8bba\u884c": idx,
                "\u60c5\u7eea": "\u597d\u8bc4" if sentiment == "positive" else "\u5dee\u8bc4",
                "\u6807\u7b7e": label,
                "\u661f\u7ea7": row.get("\u661f\u7ea7"),
                "\u547d\u4e2d\u5173\u952e\u8bcd": "\uff1b".join(f"{word}:{count}" for word, count in word_counts.items()),
                "\u547d\u4e2d\u539f\u56e0": "\u6b63\u5411\u8868\u8fbe" if sentiment == "positive" else "\u5dee\u8bc4\u8868\u8fbe\uff0c\u5df2\u6392\u9664 not/no/without/doesn't \u7b49\u5426\u5b9a\u8bed\u5883",
                "\u5173\u952e\u8bcd\u4e0a\u4e0b\u6587": "\n".join(keyword_context(text, word) for word in word_counts),
                "\u8bc1\u636e\u7247\u6bb5": text,
                "\u94fe\u63a5": row.get("\u94fe\u63a5") or "",
            })
        for label, words in SCENARIO_RULES:
            count = sum(matched_keyword_counts(text_l, words).values())
            if count:
                scenarios[label] += count
    return pos, neg, scenarios, evidence


def review_text(row: dict[str, Any]) -> str:
    return " ".join(str(row.get(k) or "") for k in ["标题", "标题 (翻译)", "内容", "内容(翻译)"]).replace("<br>", " ").strip()


def star_value(row: dict[str, Any]) -> int | None:
    value = safe_float(row.get("星级"))
    if value and 1 <= value <= 5:
        return int(value)
    return None


def matched_scenarios(text_l: str) -> list[str]:
    return [label for label, words in SCENARIO_RULES if matched_keyword_counts(text_l, words)]


def detailed_review_analysis(rows: list[dict[str, Any]]) -> dict[str, Any]:
    valid = []
    extracted_review_rows = set()
    pos = Counter()
    neg = Counter()
    pos_words = defaultdict(Counter)
    neg_words = defaultdict(Counter)
    evidence = []
    unmatched = []
    scenario_counter = Counter()
    scenario_evidence = []
    size_reviews = defaultdict(lambda: {"\u8bc4\u8bba\u6570": 0, "123\u661f\u6570\u91cf": 0, "45\u661f\u6570\u91cf": 0, "\u6570\u5b57\u7801\u6570": Counter()})
    stars = Counter()

    for idx, row in enumerate(rows, start=2):
        star = star_value(row)
        text = review_text(row)
        text_l = text.lower()
        if not star or not text_l:
            continue
        valid.append(row)
        stars[star] += 1

        model = str(row.get("\u578b\u53f7") or "")
        size_match = re.search(r"Size:\s*([^|]+)", model, re.I)
        size = (size_match.group(1).strip() if size_match else model.strip()) or "\u672a\u6807\u6ce8"
        size_reviews[size]["\u8bc4\u8bba\u6570"] += 1
        if star <= 3:
            size_reviews[size]["123\u661f\u6570\u91cf"] += 1
        else:
            size_reviews[size]["45\u661f\u6570\u91cf"] += 1
        for code in re.findall(r"\b(?:3[0-9]|4[0-4])\s*(?:A|B|C|D|DD|DDD|E|F)\b", text, flags=re.I):
            size_reviews[size]["\u6570\u5b57\u7801\u6570"][code.upper().replace(" ", "")] += 1

        hit = False
        row_scenarios = matched_scenarios(text_l)
        for scene in row_scenarios:
            scenario_counter[scene] += 1
            scenario_evidence.append({
                "\u4f7f\u7528\u573a\u666f\u6807\u7b7e": scene,
                "\u661f\u7ea7": star,
                "\u8bc1\u636e\u7247\u6bb5": text,
            })
        for label, sentiment, words in VOC_RULES:
            word_counts = matched_keyword_counts(text_l, words, sentiment)
            if not word_counts:
                continue
            hit = True
            extracted_review_rows.add(idx)
            target = pos if sentiment == "positive" else neg
            target[label] += sum(word_counts.values())
            word_counter = pos_words if sentiment == "positive" else neg_words
            for word, count in word_counts.items():
                word_counter[label][word] += count
            evidence.append({
                "\u8bc4\u8bba\u884c": idx,
                "\u60c5\u7eea": "\u597d\u8bc4" if sentiment == "positive" else "\u5dee\u8bc4",
                "\u6807\u7b7e": label,
                "\u661f\u7ea7": star,
                "\u547d\u4e2d\u5173\u952e\u8bcd": "\uff1b".join(f"{word}:{count}" for word, count in word_counts.items()),
                "\u547d\u4e2d\u539f\u56e0": "\u6b63\u5411\u8868\u8fbe" if sentiment == "positive" else "\u5dee\u8bc4\u8868\u8fbe\uff0c\u5df2\u6392\u9664 not/no/without/doesn't \u7b49\u5426\u5b9a\u8bed\u5883",
                "\u5173\u952e\u8bcd\u4e0a\u4e0b\u6587": "\n".join(keyword_context(text, word) for word in word_counts),
                "\u8bc1\u636e\u7247\u6bb5": text,
                "\u4f7f\u7528\u573a\u666f\u63d0\u53d6": "\uff1b".join(row_scenarios),
                "\u94fe\u63a5": row.get("\u94fe\u63a5") or "",
            })
        if not hit:
            unmatched.append({
                "\u661f\u7ea7": star,
                "\u8bc4\u8bba\u7247\u6bb5": text,
                "\u4f7f\u7528\u573a\u666f\u63d0\u53d6": "\uff1b".join(row_scenarios),
                "\u94fe\u63a5": row.get("\u94fe\u63a5") or "",
            })

    return {
        "valid": valid,
        "extracted_count": len(extracted_review_rows),
        "stars": stars,
        "pos": pos,
        "neg": neg,
        "pos_words": pos_words,
        "neg_words": neg_words,
        "evidence": evidence,
        "unmatched": unmatched,
        "scenario_counter": scenario_counter,
        "scenario_evidence": scenario_evidence,
        "size_reviews": size_reviews,
    }


def keyword_summary(counter: Counter, n: int = 8) -> str:
    return "；".join(f"{word}({count})" for word, count in counter.most_common(n))


def scenario_keywords(label: str) -> str:
    for scene, words in SCENARIO_RULES:
        if scene == label:
            return "；".join(words[:8])
    return ""


def outfit_pairing(label: str) -> str:
    mapping = {
        "基础内搭/外套内搭": "衬衫、开衫、西装、外套内搭",
        "吊带/背心/细肩带穿搭": "吊带、背心、细肩带上衣",
        "贴身/浅色/无痕内搭": "白色上衣、贴身T恤、浅色裙装",
        "日常/通勤/全天穿着": "通勤上衣、日常休闲装",
        "夏季/度假/旅行": "夏季短上衣、度假穿搭、旅行行李",
        "运动/瑜伽/健身": "瑜伽服、健身上衣",
        "派对/约会/外出": "约会装、派对上衣、外穿搭配",
        "居家/睡眠/放松": "居家服、睡衣、休闲装",
    }
    return mapping.get(label, "评论未明确搭配，需结合证据判断")


def kano_for_label(label: str) -> str:
    if any(key in label for key in ["尺码", "遮点", "质量", "卷边", "支撑"]):
        return "基本型/期望型"
    if any(key in label for key in ["版型", "外观", "罩杯"]):
        return "期望型"
    return "魅力型/期望型"


def deep_logic(label: str, good: int, bad: int) -> str:
    if bad > good:
        return f"{label} 已成为阻碍购买或复购的核心风险，需要优先在版型、面料或结构上验证。"
    return f"{label} 是当前好评的主要驱动点，开发时应保留并在页面卖点中放大。"


def order_stats(rows: list[dict[str, Any]]) -> tuple[float, Counter, Counter]:
    total = 0.0
    colors = Counter()
    sizes = Counter()
    for row in rows:
        qty = safe_float(row.get("近30天订单量")) or 0
        total += qty
        colors[str(row.get("Color (颜色)") or "未知")] += qty
        sizes[str(row.get("Size (尺寸)") or "未知")] += qty
    return total, colors, sizes


def top_items(counter: Counter, denominator: float | None = None, n: int = 5, unit: str = "次") -> str:
    parts = []
    for key, value in counter.most_common(n):
        if denominator:
            parts.append(f"{key}:{int(value)}{unit}({pct(value, denominator):.0%})")
        else:
            parts.append(f"{key}:{int(value)}{unit}")
    return "；".join(parts)


def expert_analysis_rows(asin: str, brand: str, pos: Counter, neg: Counter, sizes: Counter, colors: Counter, total: float) -> list[dict[str, Any]]:
    rows = []
    voc_total = sum(pos.values()) + sum(neg.values())
    core_size = top_items(sizes, total, 3, "单")
    core_color = top_items(colors, total, 3, "单")
    all_tags = set(pos.keys()) | set(neg.keys())
    ranked = sorted(all_tags, key=lambda tag: (neg.get(tag, 0) + pos.get(tag, 0), neg.get(tag, 0)), reverse=True)
    for tag in ranked[:8]:
        rule = EXPERT_RULES.get(tag, {
            "jtbd": "用户希望该体验稳定、少试错、可复购。",
            "barrier": "该点反复出现会影响外穿信心、退货和评分。",
            "mechanism": "需要结合版型、面料、杯结构和尺码映射验证。",
            "opportunity": "保留正向卖点，同时在样衣试穿中验证负向反馈。",
            "kano": "期望型",
            "effort": "中",
        })
        good = pos.get(tag, 0)
        bad = neg.get(tag, 0)
        reach = "高" if good + bad >= 20 or total >= 1000 else ("中" if good + bad >= 6 else "低")
        impact = "高" if bad and any(key in tag for key in ["支撑", "罩杯", "遮点", "尺码", "卷边"]) else ("中" if bad else "中低")
        confidence = "高" if good + bad >= 10 else "中"
        priority = "P1" if impact == "高" and confidence in {"高", "中"} else ("P2" if good + bad >= 5 else "P3")
        rows.append({
            "ASIN": asin,
            "品牌": brand,
            "VOC产品问题/机会": tag,
            "好评提及次数": int(good),
            "差评提及次数": int(bad),
            "观点占比": pct(good + bad, voc_total),
            "用户任务(JTBD)": rule["jtbd"],
            "痛点/阻碍": rule["barrier"],
            "真实情绪": "怕尴尬/怕不安全/不想反复调整" if bad else "省事、舒适、显身材带来确定感",
            "可能结构原因": rule["mechanism"],
            "开发机会": rule["opportunity"],
            "Kano属性": rule["kano"],
            "RICE-Reach": reach,
            "RICE-Impact": impact,
            "RICE-Confidence": confidence,
            "Effort": rule["effort"],
            "优先级": priority,
            "订单核心尺码": core_size,
            "订单核心颜色": core_color,
            "截图建议": "放入第一份资料对应VOC板块；若为差评Top点，同步放买家晒图/评论证据截图。",
        })
    return rows


def load_font(size: int, bold: bool = False):
    candidates = [
        r"C:\Windows\Fonts\msyhbd.ttc" if bold else r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\arial.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size)
        except Exception:
            continue
    return ImageFont.load_default()


def draw_table_image(path: Path, title: str, subtitle: str, headers: list[str], rows: list[list[Any]], widths: list[int]):
    row_h = 44
    title_h = 88
    header_h = 46
    width = sum(widths)
    height = title_h + header_h + max(5, len(rows)) * row_h + 24
    img = Image.new("RGB", (width, height), "#FFFFFF")
    draw = ImageDraw.Draw(img)
    title_font = load_font(24, True)
    head_font = load_font(16, True)
    body_font = load_font(15)
    small_font = load_font(14)
    draw.rectangle([0, 0, width, title_h], fill="#17324D")
    draw.text((24, 16), title, fill="#FFFFFF", font=title_font)
    draw.text((24, 52), subtitle, fill="#DCEBFA", font=small_font)
    y = title_h
    draw.rectangle([0, y, width, y + header_h], fill="#C45A18")
    x = 0
    for header, col_w in zip(headers, widths):
        draw.text((x + 12, y + 12), header, fill="#FFFFFF", font=head_font)
        x += col_w
    y += header_h
    for i in range(max(5, len(rows))):
        fill = "#F6FBFE" if i % 2 == 0 else "#E8F5FB"
        draw.rectangle([0, y, width, y + row_h], fill=fill)
        if i < len(rows):
            x = 0
            for value, col_w in zip(rows[i], widths):
                draw.text((x + 12, y + 11), str(value)[:34], fill="#1F2933", font=body_font)
                x += col_w
        y += row_h
    x = 0
    for col_w in widths:
        draw.line([x, title_h, x, y], fill="#CAD6E0")
        x += col_w
    draw.line([width - 1, title_h, width - 1, y], fill="#CAD6E0")
    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path)


def make_screenshots(job_dir: Path, asin: str, brand: str, pos: Counter, neg: Counter, scenarios: Counter, sizes: Counter, total: float) -> dict[str, str]:
    shot_dir = job_dir / "screenshots"
    voc_rows = []
    for label, value in pos.most_common(5):
        voc_rows.append(["好评", label, value, f"{pct(value, sum(pos.values()) + sum(neg.values())):.0%}"])
    for label, value in neg.most_common(5):
        voc_rows.append(["差评", label, value, f"{pct(value, sum(pos.values()) + sum(neg.values())):.0%}"])
    voc_path = shot_dir / f"{asin}_VOC.png"
    draw_table_image(voc_path, f"{asin} 产品VOC完整截图", brand, ["情绪", "VOC点", "提及次数", "占比"], voc_rows, [100, 420, 130, 120])

    order_path = shot_dir / f"{asin}_订单码数.png"
    order_rows = [[size, int(qty), f"{pct(qty, total):.0%}", i + 1] for i, (size, qty) in enumerate(sizes.most_common())]
    draw_table_image(order_path, f"{asin} 码数销量截图", f"{brand} 近30天订单量 {fmt_num(total)}", ["码数", "销量", "占比", "排名"], order_rows, [220, 150, 120, 100])

    scen_total = sum(scenarios.values())
    scene_path = shot_dir / f"{asin}_使用场景.png"
    scene_rows = [[label, count, f"{pct(count, scen_total):.0%}"] for label, count in scenarios.most_common(8)]
    draw_table_image(scene_path, f"{asin} 使用场景截图", brand, ["场景", "提及次数", "占比"], scene_rows, [420, 140, 120])
    return {"voc": str(voc_path), "order": str(order_path), "scenario": str(scene_path)}


def age_segment(asin: str, market: dict[str, Any], title: str) -> str:
    text = f"{market.get('分类人群') or ''}\n{market.get('用户画像') or ''}\n{title}".lower()
    if "25-38" in text or "28" in text:
        return "25-34主力，35-44+延展"
    if "v neck" in text or "v领" in text or "round" in text or "圆领" in text:
        return "35-44+"
    return "25-34"


def neckline(market: dict[str, Any], title: str) -> str:
    text = f"{market.get('三级分类') or ''} {title}".lower()
    if "v领" in text or "v neck" in text:
        return "V领"
    if "方领" in text or "square" in text:
        return "方领"
    if "cami" in text or "spaghetti" in text or "吊带" in text:
        return "吊带"
    if "圆领" in text or "scoop" in text or "crew" in text or "high neck" in text:
        return "圆领"
    return "其他"


def build_report(input_files: list[Path]) -> dict[str, Any]:
    job_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    job_dir = JOBS_DIR / job_id
    upload_dir = job_dir / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    local_files = []
    for src in input_files:
        dst = upload_dir / src.name
        if src.resolve() != dst.resolve():
            shutil.copy2(src, dst)
        local_files.append(dst)

    workbooks = {}
    asin_files = {}
    for path in local_files:
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            continue
        asin = detect_asin(path, wb)
        if "评论分析" in wb.sheetnames or "订单量分析" in wb.sheetnames:
            asin_files[asin] = path
            workbooks[asin] = wb
    market = market_rows(find_market_file(local_files), set(asin_files))

    records = []
    evidence_rows = []
    screenshot_rows = []
    order_rows = []
    size_rows = []
    scenario_rows = []
    persona_rows = []
    voc_full_rows = []
    voc_expert_rows = []
    top_good_rows = []
    top_bad_rows = []
    review_overview_rows = []
    voc_asin_rows = []
    evidence_sample_rows = []
    unmatched_rows = []
    demand_summary_rows = []
    size_order_review_rows = []
    size_order_rows = []
    color_asin_rows = []
    color_all_counter = Counter()
    demand_total_rows = []
    scenario_total_counter = Counter()
    scenario_total_rows = []
    scenario_detail_rows = []
    persona_detail_rows = []
    for asin, wb in workbooks.items():
        orders = read_sheet_dicts(wb, "订单量分析")
        reviews = read_sheet_dicts(wb, "评论分析")
        natural = read_sheet_dicts(wb, "自然词流量分析")
        m = market.get(asin, {})
        total, colors, sizes = order_stats(orders)
        analysis = detailed_review_analysis(reviews)
        pos, neg, scenarios, evidence = analysis["pos"], analysis["neg"], analysis["scenarios"], analysis["evidence"]
        title = str(m.get("商品标题") or (orders[0].get("商品标题") if orders else "") or "")
        brand = str(m.get("品牌") or "").strip() or (title.split()[0] if title else "")
        price = safe_float(m.get("价格")) or safe_float(orders[0].get("价格") if orders else None)
        rating = safe_float(m.get("评分")) or safe_float(orders[0].get("Review评分") if orders else None)
        review_count = safe_float(m.get("评论数")) or safe_float(orders[0].get("Review数量") if orders else None)
        daily_sales = safe_float(m.get("日均销量")) or (total / 30 if total else None)
        rank = first_present(m, ["小类排名", "类目排名", "BSR", "Best Sellers Rank", "排名"]) or "源表未给出"
        amz_url = product_amz_url(asin, m, orders)
        product_image = product_image_source(wb, asin, amz_url)
        buyer_images = review_image_evidence(reviews, amz_url)
        shots = make_screenshots(job_dir, asin, brand, pos, neg, scenarios, sizes, total)
        voc_total = sum(pos.values()) + sum(neg.values())
        scen_total = sum(scenarios.values())
        age = age_segment(asin, m, title)
        neck = neckline(m, title)
        top_good = top_items(pos, voc_total, 5)
        top_bad = top_items(neg, voc_total, 5)
        record = {
            "开发路径": f"{age}：{neck}修身带罩杯，按领型喜好度切入",
            "产品图片": product_image,
            "买家晒图": buyer_images,
            "产品特点": str(m.get("产品卖点") or "")[:900],
            "产品图片_2": title[:500],
            "竞对 ASIN": asin,
            "品牌/角色": f"{brand} / {age}，{neck}",
            "日均销量": daily_sales,
            "小类排名": rank,
            "上架时间": m.get("上架时间") or "源表未给出",
            "结论_1": "销量强，可作为主竞品" if daily_sales and daily_sales >= 50 else "适合作为细分参考",
            "价格($)": price,
            "结论_2": "价格带可参考" if price and price < 20 else "需用功能/面料支撑客单",
            "评分": rating,
            "结论_3": "评分表现较稳" if rating and rating >= 4.3 else "评分有优化空间",
            "评论数": review_count,
            "竞对人群定位": f"{age}；{neck}；带罩杯/修身",
            "订单": f"近30天订单约{fmt_num(total)}；颜色：{top_items(colors, total, 3, '单')}；尺码：{top_items(sizes, total, 3, '单')}",
            "结论1：用户画像确认": "基于订单核心码、领型和VOC，判断为免穿内衣、舒适显身材、一件多穿用户。",
            "VOC分析": f"好评VOC：{top_good}\n差评VOC：{top_bad}",
            "Top3好评需求(VOC占比)": top_items(pos, voc_total, 3),
            "Top5差评痛点(VOC占比)": top_bad,
            "TOP差评点占比": pct(neg.most_common(1)[0][1], voc_total) if neg else 0,
            "结论2痛点定位": f"优先解决：{neg.most_common(1)[0][0] if neg else '暂无集中差评'}；保留：{pos.most_common(1)[0][0] if pos else '暂无集中好评'}。",
            "难度系数": "中高" if neg and any(k in neg.most_common(1)[0][0] for k in ["支撑", "罩杯", "尺码"]) else "中",
            "主打词": "；".join(str(r.get("关键词 (数据来源于西柚洞察)") or "") for r in natural[:8] if r.get("关键词 (数据来源于西柚洞察)")),
            "使用场景": top_items(scenarios, scen_total, 5),
            "结论5": "可开发为主线款" if daily_sales and daily_sales >= 50 else "作为辅助验证款",
            "我方定位": f"{age}{neck}带罩杯修身背心：固定杯/轻支撑/不透/副乳包裹/可外穿",
            "优先级": "P1" if daily_sales and daily_sales >= 50 else "P2",
            "真实用户晒图来源": buyer_images,
            "带图评论链接": buyer_images,
            "AMZ链接": amz_url,
            "订单码数截图": shots["order"],
            "VOC好评差评截图": shots["voc"],
            "使用场景截图": shots["scenario"],
        }
        records.append(record)
        for size, qty in sizes.most_common():
            size_rows.append({
                "ASIN": asin,
                "品牌": brand,
                "尺码": size,
                "近30天订单量": int(qty),
                "占比": pct(qty, total),
            })
        for color, qty in colors.most_common():
            order_rows.append({
                "ASIN": asin,
                "品牌": brand,
                "颜色": color,
                "近30天订单量": int(qty),
                "占比": pct(qty, total),
            })
        for scene, count in scenarios.most_common():
            scenario_rows.append({
                "ASIN": asin,
                "品牌": brand,
                "使用场景": scene,
                "提及次数": int(count),
                "占比": pct(count, sum(scenarios.values())),
            })
        persona_rows.append({
            "ASIN": asin,
            "品牌": brand,
            "年龄段": age,
            "领型": neck,
            "核心尺码": top_items(sizes, total, 3, "单"),
            "核心颜色": top_items(colors, total, 3, "单"),
            "人物画像": record["结论1：用户画像确认"],
            "开发定位": record["我方定位"],
        })
        voc_total = sum(pos.values()) + sum(neg.values())
        for label, count in pos.most_common():
            voc_full_rows.append({"ASIN": asin, "品牌": brand, "情绪": "好评", "VOC点": label, "提及次数": int(count), "占比": pct(count, voc_total)})
        for label, count in neg.most_common():
            voc_full_rows.append({"ASIN": asin, "品牌": brand, "情绪": "差评", "VOC点": label, "提及次数": int(count), "占比": pct(count, voc_total)})
        for label, count in pos.most_common(5):
            top_good_rows.append({"ASIN": asin, "品牌": brand, "TOP5好评": label, "提及次数": int(count), "占比": pct(count, voc_total)})
        for label, count in neg.most_common(5):
            top_bad_rows.append({"ASIN": asin, "品牌": brand, "TOP5差评": label, "提及次数": int(count), "占比": pct(count, voc_total)})
        voc_expert_rows.extend(expert_analysis_rows(asin, brand, pos, neg, sizes, colors, total))
        for ev in evidence:
            evidence_rows.append([asin, brand, ev["标签"], ev["情绪"], ev["星级"], ev["证据片段"], ev["链接"]])
        for block, content, basis, source in [
            ("竞品横向分析-订单", "每个ASIN各码数销量统计截图", record["订单"], shots["order"]),
            ("竞品横向分析-VOC分析/Top3/Top5", "完整产品VOC截图：好评点、差评点、提及次数、占比", record["VOC分析"], shots["voc"]),
            ("竞品横向分析-使用场景", "评论使用场景占比截图", record["使用场景"], shots["scenario"]),
            ("竞品横向分析-买家晒图", "真实页面用户晒图或带图评论链接", record["买家晒图"], record["带图评论链接"]),
        ]:
            screenshot_rows.append([asin, brand, block, content, basis, "P1" if "买家" not in block else "P2", source])

        valid_count = len(analysis["valid"])
        star_123 = sum(analysis["stars"].get(i, 0) for i in [1, 2, 3])
        star_45 = sum(analysis["stars"].get(i, 0) for i in [4, 5])
        pos_total = sum(pos.values())
        neg_total = sum(neg.values())
        review_overview_rows.append({
            "ASIN": asin,
            "品牌": brand,
            "商品标题": title,
            "有效评论数": valid_count,
            "提取有效评论数": analysis["extracted_count"],
            "123星数量": star_123,
            "45星数量": star_45,
            "差评占比": pct(star_123, valid_count),
        })

        neg_ranked = neg.most_common()
        for idx, (good_label, good_count) in enumerate(pos.most_common()):
            bad_label, bad_count = neg_ranked[idx] if idx < len(neg_ranked) else ("", 0)
            voc_asin_rows.append({
                "ASIN": asin,
                "品牌": brand,
                "商品标题": title,
                "好评标签（中文）": good_label,
                "概括好评关键词（原文英文）": keyword_summary(analysis["pos_words"][good_label]),
                "好评频次": int(good_count),
                "好评占比": pct(good_count, pos_total),
                "对应的差评关键词（中文）": bad_label,
                "差评关键词（英文原文）": keyword_summary(analysis["neg_words"][bad_label]) if bad_label else "",
                "差评频次": int(bad_count),
                "差评占比": pct(bad_count, neg_total),
                "深度分析（逻辑）": deep_logic(good_label, good_count, bad_count),
                "Kano属性": kano_for_label(good_label),
            })

        for ev in analysis["evidence"]:
            evidence_sample_rows.append({
                "ASIN": asin,
                "品牌": brand,
                "标签": ev["标签"],
                "情绪": ev["情绪"],
                "星级": ev["星级"],
                "原文证据片段": ev["证据片段"],
                "使用场景提取": ev["使用场景提取"],
                "评论链接": ev["链接"],
            })
        for item in analysis["unmatched"]:
            unmatched_rows.append({
                "ASIN": asin,
                "品牌": brand,
                "星级": item["星级"],
                "未归类评论片段": item["评论片段"],
                "使用场景提取": item["使用场景提取"],
                "评论链接": item["链接"],
            })

        demand_summary = {
            "ASIN": asin,
            "品牌": brand,
            "有效评论数": valid_count,
            "45星评论数": star_45,
            "45星评论占比": pct(star_45, valid_count),
            "123星评论数": star_123,
            "123星评论占比": pct(star_123, valid_count),
            "好评观点总频次": pos_total,
            "差评观点总频次": neg_total,
            "Top好评需求层级": pos.most_common(1)[0][0] if pos else "",
            "Top核心差评点": neg.most_common(1)[0][0] if neg else "",
        }
        demand_summary_rows.append(demand_summary)

        for size, qty in sizes.most_common():
            review_size = analysis["size_reviews"].get(size, {})
            review_size_count = review_size.get("评论数", 0) if review_size else 0
            bad_size = review_size.get("123星数量", 0) if review_size else 0
            good_size = review_size.get("45星数量", 0) if review_size else 0
            size_codes = review_size.get("数字码数", Counter()) if review_size else Counter()
            size_order_rows.append({
                "ASIN": asin,
                "商品标题": title,
                "码数": size,
                "完整尺码": "",
                "订单量": int(qty),
                "订单量占比": pct(qty, total),
            })
            size_order_review_rows.append({
                **demand_summary,
                "码数": size,
                "完整尺码": "",
                "订单量": int(qty),
                "订单量占比": pct(qty, total),
                "该码数评论数": review_size_count,
                "123星数量": bad_size,
                "123星占比": pct(bad_size, valid_count),
                "45星数量": good_size,
                "45星占比": pct(good_size, valid_count),
                "评论正文数字码数提及": keyword_summary(size_codes),
                "数字码数映射依据": "评论型号 Size 字段 + 评论正文数字杯码识别；完整尺码需由尺码图/页面尺码表补全",
            })

        for color, qty in colors.most_common():
            color_asin_rows.append({
                "ASIN": asin,
                "商品标题": title,
                "颜色名称": color,
                "订单量": int(qty),
                "订单量占比": pct(qty, total),
            })
            color_all_counter[color] += qty

        for label, count in pos.items():
            demand_total_rows.append({"需求标签": label, "情绪": "好评", "提及次数": int(count), "ASIN": asin, "品牌": brand})
        for label, count in neg.items():
            demand_total_rows.append({"需求标签": label, "情绪": "差评", "提及次数": int(count), "ASIN": asin, "品牌": brand})

        scenario_total_counter.update(scenarios)
        scen_total_for_asin = sum(scenarios.values())
        for scene, count in scenarios.most_common():
            scenario_detail_rows.append({
                "ASIN": asin,
                "品牌": brand,
                "使用场景标签": scene,
                "代表关键词": scenario_keywords(scene),
                "服装搭配": outfit_pairing(scene),
                "提及次数": int(count),
                "占比": pct(count, scen_total_for_asin),
            })
        persona_detail_rows.append({
            "ASIN": asin,
            "品牌": brand,
            "商品标题": title,
            "用户画像": record["结论1：用户画像确认"],
            "核心问题": demand_summary["Top核心差评点"],
            "核心需求": demand_summary["Top好评需求层级"],
            "核心尺码": top_items(sizes, total, 3, "单"),
            "核心颜色": top_items(colors, total, 3, "单"),
            "使用场景": top_items(scenarios, scen_total_for_asin, 5),
            "开发建议": record["我方定位"],
        })

    color_all_total = sum(color_all_counter.values())
    color_all_rows = [
        {
            "颜色名称": color,
            "总订单量": int(qty),
            "总订单量占比": pct(qty, color_all_total),
        }
        for color, qty in color_all_counter.most_common()
    ]
    demand_counter = Counter()
    demand_good = Counter()
    demand_bad = Counter()
    for row in demand_total_rows:
        demand_counter[row["需求标签"]] += row["提及次数"]
        if row["情绪"] == "好评":
            demand_good[row["需求标签"]] += row["提及次数"]
        else:
            demand_bad[row["需求标签"]] += row["提及次数"]
    demand_total = sum(demand_counter.values())
    demand_summary_all_rows = [
        {
            "需求标签": label,
            "总提及次数": int(count),
            "总占比": pct(count, demand_total),
            "好评提及次数": int(demand_good.get(label, 0)),
            "差评提及次数": int(demand_bad.get(label, 0)),
            "Kano属性": kano_for_label(label),
            "开发判断": deep_logic(label, demand_good.get(label, 0), demand_bad.get(label, 0)),
        }
        for label, count in demand_counter.most_common()
    ]
    scenario_total = sum(scenario_total_counter.values())
    scenario_total_rows = [
        {
            "ASIN": "全部ASIN汇总",
            "品牌": "全部竞对",
            "使用场景标签": scene,
            "代表关键词": scenario_keywords(scene),
            "服装搭配": outfit_pairing(scene),
            "提及次数": int(count),
            "占比": pct(count, scenario_total),
        }
        for scene, count in scenario_total_counter.most_common()
    ]

    module_data = {
        "01_竞对好评差评总览": review_overview_rows,
        "02_VOC全量分析_ASIN好评差评": voc_asin_rows,
        "03_评论证据样例": evidence_sample_rows,
        "04_未归类观点检查": unmatched_rows,
        "05_各ASIN需求层级摘要": demand_summary_rows,
        "06_ASIN码数订单评论": size_order_review_rows,
        "07_尺码订单统计表": size_order_rows,
        "08_ASIN颜色订单占比": color_asin_rows,
        "09_颜色统计表": color_all_rows,
        "10_需求汇总统计": demand_summary_all_rows,
        "11_使用场景汇总": scenario_total_rows,
        "12_各ASIN使用场景明细表": scenario_detail_rows,
        "13_用户画像": persona_detail_rows,
        "14_竞对横向数据对比表": records,
        "TOP5好评": top_good_rows,
        "TOP5差评": top_bad_rows,
        "截图清单": [
            {
                "ASIN": r[0],
                "品牌": r[1],
                "第一份资料原板块": r[2],
                "需要截图内容": r[3],
                "VOC依据": r[4],
                "截图优先级": r[5],
                "现有截图文件/来源": r[6],
            }
            for r in screenshot_rows
        ],
    }
    output = job_dir / "竞品横向分析自动生成报告.xlsx"
    export_workbook(output, records, evidence_rows, screenshot_rows, module_data)
    module_downloads = export_module_workbooks(job_dir, module_data)
    preview = {
        "jobId": job_id,
        "downloadUrl": f"/download/{job_id}/竞品横向分析自动生成报告.xlsx",
        "asinCount": len(records),
        "records": records,
        "directions": [
            "方向1：35-44+ V领/圆领带罩杯通勤款，主打免穿内衣、舒适、不透。",
            "方向2：25-34 方领带罩杯外穿款，主打显身材、稳定杯位、浅色安全。",
            "方向3：25-34 吊带/可调肩带场景款，主打夏季、旅行、约会和一件多穿。",
        ],
        "modules": list(module_data.keys()),
        "moduleData": module_data,
        "moduleDownloads": module_downloads,
    }
    (job_dir / "preview.json").write_text(json.dumps(preview, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return preview


def cell_value(record: dict[str, Any], header: str, idx: int) -> Any:
    duplicate = {
        4: "产品图片_2",
        10: "结论_1",
        12: "结论_2",
        14: "结论_3",
    }
    key = duplicate.get(idx, header)
    if header.endswith("截图"):
        return "见截图页"
    return record.get(key, "")


def horizontal_row(record: dict[str, Any]) -> dict[str, Any]:
    return {header: cell_value(record, header, idx) for idx, header in enumerate(FIELDS)}


def style_sheet(ws, header_fill="FFC45A18"):
    orange = PatternFill("solid", fgColor=header_fill)
    white = Font(color="FFFFFF", bold=True, name="Microsoft YaHei", size=10)
    body = Font(name="Microsoft YaHei", size=10, color="1F2933")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = orange
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.auto_filter.ref = ws.dimensions


def append_dict_sheet(wb, title: str, rows: list[dict[str, Any]]):
    ws = wb.create_sheet(title[:31])
    if not rows:
        ws.append(["暂无数据"])
        style_sheet(ws)
        return ws
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_sheet(ws)
    for idx, header in enumerate(headers, start=1):
        width = 18
        if any(k in str(header) for k in ["VOC", "画像", "定位", "依据", "链接", "产品", "场景"]):
            width = 42
        ws.column_dimensions[get_column_letter(idx)].width = width
    return ws


def export_module_workbooks(job_dir: Path, module_data: dict[str, list[dict[str, Any]]]) -> dict[str, str]:
    downloads = {}
    for name, rows in module_data.items():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = name[:31]
        if name in {"竞对横向分析", "14_竞对横向数据对比表"} and rows:
            ws.append(FIELDS)
            for row in rows:
                ws.append([cell_value(row, header, idx) for idx, header in enumerate(FIELDS)])
        elif rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h, "") for h in headers])
        else:
            ws.append(["暂无数据"])
        style_sheet(ws)
        for idx, header in enumerate(ws[1], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = 28
        filename = f"{name}.xlsx"
        wb.save(job_dir / filename)
        downloads[name] = f"/download/{job_dir.name}/{filename}"
    return downloads


def export_workbook(path: Path, records: list[dict[str, Any]], evidence_rows: list[list[Any]], screenshot_rows: list[list[Any]], module_data: dict[str, list[dict[str, Any]]]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "14_竞对横向数据对比表"
    ws.append(FIELDS)
    for record in records:
        ws.append([cell_value(record, header, i) for i, header in enumerate(FIELDS)])
    style_sheet(ws)
    widths = [22, 42, 58, 48, 38, 15, 24, 12, 12, 13, 18, 10, 18, 9, 16, 10, 28, 46, 42, 56, 42, 48, 14, 42, 10, 38, 40, 18, 42, 10]
    for i, width in enumerate(widths[:len(FIELDS)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 120

    summary = wb.create_sheet("开发结论")
    summary.append(["模块", "结论"])
    for row in [
        ["总判断", "优先做带罩杯修身背心，拆成25-34外穿显身材、35-44+通勤日常舒适升级两条路径。"],
        ["开发方向1", "35-44+ V领/圆领带罩杯通勤款：轻支撑固定杯、双层不透、低存在感、基础色。"],
        ["开发方向2", "25-34 方领带罩杯外穿款：低方领/修身显腰、稳定杯位、副乳包裹、浅色不露杯。"],
        ["开发方向3", "25-34 吊带/可调肩带场景款：细肩带/可拆肩带，可内搭可单穿，夏季旅行约会补充。"],
    ]:
        summary.append(row)
    style_sheet(summary)
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 100

    shot = wb.create_sheet("订单VOC场景截图")
    shot.append(["ASIN", "品牌/角色", "订单码数销量截图", "VOC好评差评截图", "使用场景截图"])
    style_sheet(shot)
    shot.column_dimensions["A"].width = 16
    shot.column_dimensions["B"].width = 28
    for col in "CDE":
        shot.column_dimensions[col].width = 38
    for i, record in enumerate(records, start=2):
        shot.cell(i, 1).value = record["竞对 ASIN"]
        shot.cell(i, 2).value = record["品牌/角色"]
        shot.row_dimensions[i].height = 190
        for col, key in [(3, "订单码数截图"), (4, "VOC好评差评截图"), (5, "使用场景截图")]:
            image_path = record.get(key)
            if image_path and Path(image_path).exists():
                img = XlImage(image_path)
                img.width = 260
                img.height = 165
                shot.add_image(img, f"{get_column_letter(col)}{i}")

    need = wb.create_sheet("第一份资料截图清单")
    need.append(["ASIN", "品牌", "第一份资料原板块", "需要截图内容", "VOC依据", "截图优先级", "现有截图文件/来源"])
    for row in screenshot_rows:
        need.append(row)
    style_sheet(need)
    for col, width in zip("ABCDEFG", [16, 18, 28, 36, 60, 12, 60]):
        need.column_dimensions[col].width = width

    ev = wb.create_sheet("VOC证据片段")
    ev.append(["ASIN", "品牌", "VOC标签", "情绪", "星级", "证据片段", "评论链接"])
    for row in evidence_rows:
        ev.append(row)
    style_sheet(ev)
    for col, width in zip("ABCDEFGHIJ", [16, 18, 28, 10, 10, 28, 36, 70, 110, 60]):
        ev.column_dimensions[col].width = width

    for name, rows in module_data.items():
        if name in {"竞对横向分析", "14_竞对横向数据对比表", "截图清单"}:
            continue
        append_dict_sheet(wb, name, rows)

    preferred = [
        "01_竞对好评差评总览",
        "02_VOC全量分析_ASIN好评差评",
        "03_评论证据样例",
        "04_未归类观点检查",
        "05_各ASIN需求层级摘要",
        "06_ASIN码数订单评论",
        "07_尺码订单统计表",
        "08_ASIN颜色订单占比",
        "09_颜色统计表",
        "10_需求汇总统计",
        "11_使用场景汇总",
        "12_各ASIN使用场景明细表",
        "13_用户画像",
        "14_竞对横向数据对比表",
    ]
    wb._sheets = sorted(wb._sheets, key=lambda sheet: preferred.index(sheet.title) if sheet.title in preferred else len(preferred))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
